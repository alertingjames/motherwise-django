# import datetime
import difflib
import os
import string
import urllib
from itertools import islice

import io
import requests
import xlrd
import re

from django.core import mail
from django.core.mail import send_mail, BadHeaderError, EmailMessage
from django.contrib import messages
# from _mysql_exceptions import DataError, IntegrityError
from django.template import RequestContext

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives

from django.core.files.storage import FileSystemStorage
import json
from django.contrib import auth
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, HttpResponseNotAllowed
from django.utils.datastructures import MultiValueDictKeyError
from django.views.decorators.cache import cache_control
from numpy import long
from openpyxl.styles import PatternFill

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.fields import empty
from rest_framework.permissions import AllowAny
from time import gmtime, strftime
import time
from xlrd import XLRDError

from django.db.models import Q

from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib.sessions.backends.db import SessionStore
from django.contrib.sessions.models import Session
from django.contrib.auth.models import User, AnonymousUser
from django.conf import settings
from django import forms
import sys
from django.core.cache import cache
import random
import emoji

from linkpreview import link_preview
import favicon
import certifi
import ssl
from bs4 import BeautifulSoup
from urllib.request import Request, urlopen
from selenium import webdriver

from pbsonesignal import PybossaOneSignal

from pyfcm import FCMNotification

from motherwise.models import Member, Contact, Group, GroupMember, GroupConnect, Post, PostUrlPreview, Comment, PostPicture, PostLike, Notification, Received, Sent, Replied, Conference, Report, PostCategory
from motherwise.models import Cohort, CommentLike
from motherwise.serializers import PostSerializer, CommentSerializer

import pyrebase

config = {
    "apiKey": "AIzaSyDveaXbV1HHMREyZzbvQe53BJEHVIRCf14",
    "authDomain": "motherwise-1585202524394.firebaseapp.com",
    "databaseURL": "https://motherwise-1585202524394.firebaseio.com",
    "storageBucket": "motherwise-1585202524394.appspot.com"
}

firebase = pyrebase.initialize_app(config)


class UploadFileForm(forms.Form):
    file = forms.FileField()

def index(request):
    # try:
    #     if request.session['adminID'] != 0:
    #         return redirect('/manager/home')
    # except KeyError:
    #     print('no session')

    try:
        if request.session['memberID'] != '' and request.session['memberID'] != 0:
            member_id = request.session['memberID']
            members = Member.objects.filter(id=member_id)
            if members.count() == 0:
                return render(request, 'mothers/login.html')
            member = members[0]
            if member.cohort == 'admin':
                return render(request, 'mothers/login.html')
            if member.photo_url == '' or member.cohort == '' or member.phone_number == '':
                return render(request, 'mothers/register_profile.html', {'member':member})
            elif member.address == '' or member.city == '':
                return  render(request, 'mothers/location_picker.html', {'address':member.address})
            else:
                return redirect('/mothers/zzzzz')
    except KeyError:
        print('no session')
    return render(request, 'mothers/login.html')

    # return redirect('/mothers/')




def admin(request):
    # return redirect('/manager/logout')
    try:
        if request.session['adminID'] != '' and request.session['adminID'] != 0:
            return redirect('/manager/home')
    except KeyError:
        print('no session')
    return render(request, 'motherwise/admin.html')

def adminsignuppage(request):
    return render(request, 'motherwise/adminsignup.html')

def adminloginpage(request):
    return render(request, 'motherwise/admin.html')


@api_view(['GET', 'POST'])
def adminSignup(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        name = request.POST.get('name', '')
        password = request.POST.get('password', '')
        phone_number = request.POST.get('phone_number', '')
        playerID = request.POST.get('playerID', '')

        members = Member.objects.filter(email=email, cohort='admin')
        count = members.count()
        if count ==0:
            member = Member()
            member.admin_id = '0'
            member.email = email
            # member.name = name
            member.name = 'MotherWise Admin'
            member.password = password
            member.phone_number = phone_number
            member.photo_url = settings.URL + '/static/images/manager.jpg'
            member.cohort = 'admin'
            member.registered_time = str(int(round(time.time() * 1000)))
            member.playerID = playerID
            member.save()

            request.session['adminID'] = member.pk

            return redirect('/manager/home')

        else:
            return redirect('/manager/logout')


def adminhome(request):
    import datetime
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    members = Member.objects.filter(admin_id=adminID).order_by('-id')
    for member in members:
        if member.registered_time != '':
            member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")

    users = members[:25]
    first_page = 1
    last_page = int(members.count() / 25)
    if members.count() % 25 > 0: last_page += 1
    s = 1
    e = 7
    if e > last_page: e = last_page

    groups = Group.objects.filter(member_id=admin.pk).order_by('-id')

    return render(request, 'motherwise/adminhome2.html', {'me':admin, 'users':users, 'range':range(s, e + 1), 'current':'1', 'groups':groups, 'cohorts':cohorts, 'last_page':str(last_page)})


def get_all_member_data(members):
    memberList = members[:25]
    first_page = 1
    last_page = int(members.count() / 25)
    if members.count() % 25 > 0: last_page += 1
    s = 1
    e = 7
    if e > last_page: e = last_page

    return memberList, range(s, e + 1), last_page



def adminlogout(request):
    request.session['adminID'] = 0
    request.session['selected_option'] = ''
    request.session['selected_member_list'] = []
    return render(request, 'motherwise/admin.html')



@api_view(['GET', 'POST'])
def adminLogin(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        playerID = request.POST.get('playerID', '')

        members = Member.objects.filter(email=email, password=password, cohort='admin')
        if members.count() > 0:
            member = members[0]
            if playerID != '':
                member.playerID = playerID
                member.save()
            request.session['adminID'] = member.pk
            return redirect('/manager/home')
        else:
            return render(request, 'motherwise/result.html',
                          {'response': 'You don\'t have any permission to access this site. Try again with another credential.'})

def export_xlsx_member(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=member_template.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    row_num = 0

    columns = [
        (u"Name", 30),
        (u"E-mail", 50),
        (u"Phone Number (+x xxx xxx xxxx)", 30),
        (u"Group Name(Refer to Sheet2)", 30),
        (u"City Name", 30),
        (u"Address", 80),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    ws2=wb.create_sheet(title='Sheet2')

    ws2.column_dimensions["A"].width = 20
    my_color = openpyxl.styles.colors.Color(rgb='00ffaa02')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

    ws2['A1'].fill = my_fill
    ws2['A1'] = 'Group Names'
    ws2['A2'] = 'E81'
    ws2['A3'] = 'E83'
    ws2['A4'] = 'E84'
    ws2['A5'] = 'E86'
    ws2['A6'] = 'E87'
    ws2['A7'] = 'S82'
    ws2['A8'] = 'S85'
    ws2['A9'] = 'S88'
    ws2['A10'] = 'E(v)89'
    ws2['A11'] = 'E(v)90'
    ws2['A12'] = 'S(v)91'
    ws2['A13'] = 'E(v)92'
    ws2['A14'] = 'E(v)93'
    ws2['A15'] = 'E(v)94'
    ws2['A16'] = 'S(v)95'
    ws2['A17'] = 'E(v)96'
    ws2['A18'] = 'E(v)97'
    ws2['A19'] = 'S(v)98'
    ws2['A20'] = 'E(v)99'
    ws2['A21'] = 'E(v)100'
    ws2['A22'] = 'S(v)101'
    ws2['A23'] = 'MotherWise Alumni'
    ws2['A24'] = 'MotherWise Team'

    wb.save(response)
    return response


def import_view_member(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()

    return render(
        request,
        'motherwise/upload_form_member.html',
        {
            'form': form,
            'title': 'Load Data',
            'header': 'Upload Members From File'
        })


def import_member_data(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']

        if form.is_valid():
            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                sheet = book.sheet_by_index(0)

                for r in range(1, sheet.nrows):
                    name = sheet.cell(r, 0).value
                    email = sheet.cell(r, 1).value
                    phone_number = sheet.cell(r, 2).value
                    group_name = sheet.cell(r, 3).value
                    city = sheet.cell(r, 4).value
                    address = sheet.cell(r, 5).value

                    name = name.strip()
                    email = email.strip()
                    city = city.strip()
                    address = address.strip()

                    members = Member.objects.filter(email=email)
                    if members.count() > 0:
                        continue
                    member = Member()
                    member.admin_id = adminID
                    member.name = name
                    member.email = email
                    member.password = 'mama'
                    member.phone_number = str(phone_number).replace('.0','').strip()
                    # group_number = str(group_number).replace('.0','')
                    # try:
                    #   val = int(group_number)
                    #   member.cohort = 'Group-' + str(val)
                    # except ValueError:
                    #   print("That's not an int!")
                    member.cohort = str(group_name).replace('.0','').strip()
                    member.city = city
                    member.address = address
                    member.lat = '0'
                    member.lng = '0'
                    member.save()

                    admin = Member.objects.get(id=adminID)

                    groupText = ''
                    if member.cohort != '':
                        groupText = '<br>Group: ' + member.cohort

                    title = 'Invitation for MotherWise Community: The Nest'
                    subject = 'MotherWise Community: The Nest'
                    message = 'Dear ' + member.name + ',<br><br>Welcome to \"The Nest\": MotherWise\'s virtual community!<br><br>The Nest is an opportunity to connect and reconnect with other MotherWise families.<br>'
                    message = message + 'You can post articles, share pregnancy and new baby tips, watch videos, and chat directly with other moms. You\'ll also stay up-to-date on all the new programs and special events MotherWise has to offer!<br><br>'
                    message = message + settings.URL + '/nest/mothers' + '<br><br>We are providing you with your initial login information as follows:<br><br>'
                    message = message + 'E-mail: ' + member.email + ' (your email)<br><div style="color:white;background:black;padding:6px 10px 6px 10px;text-align:center;display:inline-block;margin:8px 0px 8px 0px;">Password: mama</div>' + groupText + '<br><br>'

                    message = message + '***By signing up to The Nest, you are agreeing to not engage in any type of: ***<br>'
                    message = message + '        · hate speech<br>'
                    message = message + '        · cyberbullying<br>'
                    message = message + '        · solicitation and/or selling of goods or services<br>'
                    message = message + '        · posting content inappropriate for our diverse community including but not limited to political<br>'
                    message = message + '    or religious views<br><br>'
                    message = message + 'We want The Nest to be a safe place for support and inspiration. Help us foster this community and please respect everyone on The Nest.<br><br>'
                    message = message + 'Please watch this video to see how to login: https://vimeo.com/430742850<br><br>'
                    message = message + 'If you have any question, please contact us:<br><br>'

                    message = message + '   E-mail: ' + 'motherwisecolorado@gmail.com' + '<br>   Phone number: ' + '720-504-4624<br><br>'
                    message = message + '<a href=\'' + settings.URL + '/nest/mothers' + '\' target=\'_blank\'>Join website</a><br><br>'
                    message = message + 'You can download and install the MotherWise mobile apps here:<br><br>'
                    message = message + '<a href="https://play.google.com/store/apps/details?id=com.app.motherwise"><img src="https://www.vacay.company/static/images/playstore.png" style="width:150px;"></a>' + '<br>'
                    message = message + '<a href="https://apps.apple.com/us/app/id1530809402#?platform=iphone"><img src="https://www.vacay.company/static/images/appstore.png" style="width:150px;"></a>' + '<br><br>'
                    message = message + 'Sincerely<br><br>MotherWise Team'

                    message = message + '<br><br>'

                    groupText2 = ''
                    if member.cohort != '':
                        groupText2 = '<br>Grupo: ' + member.cohort

                    title2 = 'Invitación para la comunidad de MotherWise: El Nido'
                    message2 = 'Querida ' + member.name + ',<br><br>¡Bienvenida al \"Nido\": la comunidad virtual de MotherWise!<br><br>El Nido es una oportunidad para conectarse y reconectarse con otras madres de MotherWise.<br>'
                    message2 = message2 + 'Puede publicar artículos, compartir consejos sobre embarazo y nuevos bebés, ver videos y chatear directamente con otras madres. ¡También se mantendrá al tanto sobre todos los nuevos programas y eventos especiales que MotherWise tiene para ofrecer!<br><br>'
                    message2 = message2 + settings.URL + '/nest/mothers' + '<br><br>Le proporcionamos su información de inicio de la siguiente manera:<br><br>'
                    message2 = message2 + 'Correo electrónico: ' + member.email + ' (Tu correo electrónico)<br><div style="color:white;background:black;padding:6px 10px 6px 10px;text-align:center;display:inline-block;margin:8px 0px 8px 0px;">Contraseña: mama</div>' + groupText2 + '<br><br>'

                    message2 = message2 + '***Al suscribirse al Nido, acepta no participar en ningún tipo de: ***<br>'
                    message2 = message2 + '        · El discurso del odio<br>'
                    message2 = message2 + '        · Ciberacoso<br>'
                    message2 = message2 + '        · Solicitud y/o venta de bienes o servicios<br>'
                    message2 = message2 + '        · Publicar contenido inapropiado para nuestra diversa comunidad, incluidos, entre otros, puntos de vista políticos o religiosos<br><br>'
                    message2 = message2 + 'Queremos que El Nido sea un lugar seguro para apoyo e inspiración. Por favor ayúdanos a fomentar esta comunidad y por favor respeta a todos en El Nido.<br><br>'
                    message2 = message2 + 'Mire este video para ver cómo iniciar sesión: https://vimeo.com/430742850<br><br>'
                    message2 = message2 + 'Si usted tiene cualquier pregunta, por favor póngase en contacto con nosotros:<br><br>'

                    message2 = message2 + '   Correo electrónico: ' + 'motherwisecolorado@gmail.com' + '<br>   Número de teléfono: ' + '720-504-4624<br><br>'
                    message2 = message2 + '<a href=\'' + settings.URL + '/nest/mothers' + '\' target=\'_blank\'>Unirse al sitio web</a><br><br>'
                    message2 = message2 + 'puede descargar e instalar las aplicaciones móviles de MotherWise aquí:<br><br>'
                    message2 = message2 + '<a href="https://play.google.com/store/apps/details?id=com.app.motherwise"><img src="https://www.vacay.company/static/images/playstore.png" style="width:150px;"></a>' + '<br>'
                    message2 = message2 + '<a href="https://apps.apple.com/us/app/id1530809402#?platform=iphone"><img src="https://www.vacay.company/static/images/appstore.png" style="width:150px;"></a>' + '<br><br>'
                    message2 = message2 + 'Sinceramente,<br><br>el Equipo de MotherWise'

                    from_email = admin.email
                    to_emails = []
                    to_emails.append(member.email)
                    send_mail_message0(from_email, to_emails, title, subject, message, title2, message2)

                return redirect('/manager/home')

            except XLRDError:
                return render(request, 'motherwise/upload_form_member.html', {'note': 'invalid_file'})

            except IOError:
                return render(request, 'motherwise/upload_form_member.html', {'note': 'invalid_file'})
            except IndexError:
                return render(request, 'motherwise/upload_form_member.html', {'note': 'invalid_file'})
            # except DataError:
            #     return HttpResponse('Invalid file!')
        else:
            return render(request, 'motherwise/upload_form_member.html', {'note': 'invalid_file'})



@api_view(['GET', 'POST'])
def add_member(request):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        phone_number = request.POST.get('phone_number', '')
        cohort = request.POST.get('cohort', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']

        members = Member.objects.filter(email=email)
        if members.count() > 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'This member has already been registered.'})
        member = Member()
        member.admin_id = adminID
        member.name = name
        member.email = email
        member.password = 'mama'
        member.phone_number = str(phone_number).replace('.0','')
        member.cohort = cohort
        member.lat = '0'
        member.lng = '0'
        member.save()

        admin = Member.objects.get(id=adminID)

        groupText = ''
        if member.cohort != '':
            groupText = '<br>Group: ' + member.cohort

        title = 'Invitation for MotherWise Community: The Nest'
        subject = 'MotherWise Community: The Nest'
        message = 'Dear ' + member.name + ',<br><br>Welcome to \"The Nest\": MotherWise\'s virtual community!<br><br>The Nest is an opportunity to connect and reconnect with other MotherWise families.<br>'
        message = message + 'You can post articles, share pregnancy and new baby tips, watch videos, and chat directly with other moms. You\'ll also stay up-to-date on all the new programs and special events MotherWise has to offer!<br><br>'
        message = message + settings.URL + '/nest/mothers' + '<br><br>We are providing you with your initial login information as follows:<br><br>'
        message = message + 'E-mail: ' + member.email + ' (your email)<br><div style="color:white;background:black;padding:6px 10px 6px 10px;text-align:center;display:inline-block;margin:8px 0px 8px 0px;">Password: mama</div>' + groupText + '<br><br>'

        message = message + '***By signing up to The Nest, you are agreeing to not engage in any type of: ***<br>'
        message = message + '        · hate speech<br>'
        message = message + '        · cyberbullying<br>'
        message = message + '        · solicitation and/or selling of goods or services<br>'
        message = message + '        · posting content inappropriate for our diverse community including but not limited to political<br>'
        message = message + '    or religious views<br><br>'
        message = message + 'We want The Nest to be a safe place for support and inspiration. Help us foster this community and please respect everyone on The Nest.<br><br>'
        message = message + 'Please watch this video to see how to login: https://vimeo.com/430742850<br><br>'
        message = message + 'If you have any question, please contact us:<br><br>'

        message = message + '   E-mail: ' + 'motherwisecolorado@gmail.com' + '<br>   Phone number: ' + '720-504-4624<br><br>'
        message = message + '<a href=\'' + settings.URL + '/nest/mothers' + '\' target=\'_blank\'>Join website</a><br><br>'
        message = message + 'You can download and install the MotherWise mobile apps here:<br><br>'
        message = message + '<a href="https://play.google.com/store/apps/details?id=com.app.motherwise"><img src="https://www.vacay.company/static/images/playstore.png" style="width:150px;"></a>' + '<br>'
        message = message + '<a href="https://apps.apple.com/us/app/id1530809402#?platform=iphone"><img src="https://www.vacay.company/static/images/appstore.png" style="width:150px;"></a>' + '<br><br>'
        message = message + 'Sincerely<br><br>MotherWise Team'

        message = message + '<br><br>'

        groupText2 = ''
        if member.cohort != '':
            groupText2 = '<br>Grupo: ' + member.cohort

        title2 = 'Invitación para la comunidad de MotherWise: El Nido'
        message2 = 'Querida ' + member.name + ',<br><br>¡Bienvenida al \"Nido\": la comunidad virtual de MotherWise!<br><br>El Nido es una oportunidad para conectarse y reconectarse con otras madres de MotherWise.<br>'
        message2 = message2 + 'Puede publicar artículos, compartir consejos sobre embarazo y nuevos bebés, ver videos y chatear directamente con otras madres. ¡También se mantendrá al tanto sobre todos los nuevos programas y eventos especiales que MotherWise tiene para ofrecer!<br><br>'
        message2 = message2 + settings.URL + '/nest/mothers' + '<br><br>Le proporcionamos su información de inicio de la siguiente manera:<br><br>'
        message2 = message2 + 'Correo electrónico: ' + member.email + ' (Tu correo electrónico)<br><div style="color:white;background:black;padding:6px 10px 6px 10px;text-align:center;display:inline-block;margin:8px 0px 8px 0px;">Contraseña: mama</div>' + groupText2 + '<br><br>'

        message2 = message2 + '***Al suscribirse al Nido, acepta no participar en ningún tipo de: ***<br>'
        message2 = message2 + '        · El discurso del odio<br>'
        message2 = message2 + '        · Ciberacoso<br>'
        message2 = message2 + '        · Solicitud y/o venta de bienes o servicios<br>'
        message2 = message2 + '        · Publicar contenido inapropiado para nuestra diversa comunidad, incluidos, entre otros, puntos de vista políticos o religiosos<br><br>'
        message2 = message2 + 'Queremos que El Nido sea un lugar seguro para apoyo e inspiración. Por favor ayúdanos a fomentar esta comunidad y por favor respeta a todos en El Nido.<br><br>'
        message2 = message2 + 'Mire este video para ver cómo iniciar sesión: https://vimeo.com/430742850<br><br>'
        message2 = message2 + 'Si usted tiene cualquier pregunta, por favor póngase en contacto con nosotros:<br><br>'

        message2 = message2 + '   Correo electrónico: ' + 'motherwisecolorado@gmail.com' + '<br>   Número de teléfono: ' + '720-504-4624<br><br>'
        message2 = message2 + '<a href=\'' + settings.URL + '/nest/mothers' + '\' target=\'_blank\'>Unirse al sitio web</a><br><br>'
        message2 = message2 + 'puede descargar e instalar las aplicaciones móviles de MotherWise aquí:<br><br>'
        message2 = message2 + '<a href="https://play.google.com/store/apps/details?id=com.app.motherwise"><img src="https://www.vacay.company/static/images/playstore.png" style="width:150px;"></a>' + '<br>'
        message2 = message2 + '<a href="https://apps.apple.com/us/app/id1530809402#?platform=iphone"><img src="https://www.vacay.company/static/images/appstore.png" style="width:150px;"></a>' + '<br><br>'
        message2 = message2 + 'Sinceramente,<br><br>el Equipo de MotherWise'

        from_email = admin.email
        to_emails = []
        to_emails.append(member.email)
        send_mail_message0(from_email, to_emails, title, subject, message, title2, message2)

        return redirect('/manager/home')


def send_mail_message0(from_email, to_emails, title, subject, message, title2, message2):
    html =  """\
                <html>
                    <head></head>
                    <body>
                        <a href="#"><img src="https://www.vacay.company/static/images/logo.png" style="width:120px;height:120px;border-radius: 8%; margin-left:25px;"/></a>
                        <h2 style="margin-left:10px; color:#02839a;">{title}</h2>
                        <div style="font-size:14px; white-space: pre-line; word-wrap: break-word;">
                            {mes}
                        </div>
                        <h2 style="margin-left:10px; color:#02839a;">{title2}</h2>
                        <div style="font-size:14px; white-space: pre-line; word-wrap: break-word;">
                            {mes2}
                        </div>
                    </body>
                </html>
            """
    html = html.format(title=title, mes=message, title2=title2, mes2=message2)

    msg = EmailMultiAlternatives(subject, '', from_email, to_emails)
    msg.attach_alternative(html, "text/html")
    msg.send(fail_silently=False)



def send_mail_message(from_email, to_emails, title, subject, message):
    html =  """\
                <html>
                    <head></head>
                    <body>
                        <a href="#"><img src="https://www.vacay.company/static/images/logo.png" style="width:120px;height:120px;border-radius: 8%; margin-left:25px;"/></a>
                        <h2 style="margin-left:10px; color:#02839a;">{title}</h2>
                        <div style="font-size:14px; white-space: pre-line; word-wrap: break-word;">
                            {mes}
                        </div>
                    </body>
                </html>
            """
    html = html.format(title=title, mes=message)

    msg = EmailMultiAlternatives(subject, '', from_email, to_emails)
    msg.attach_alternative(html, "text/html")
    msg.send(fail_silently=False)



def generateRandomPassword():
    import strgen
    randomString = strgen.StringGenerator("[\w\d]{10}").render()
    return randomString




@api_view(['GET', 'POST'])
def delete_member(request):
    if request.method == 'GET':
        member_id = request.GET['member_id']

        members = Member.objects.filter(id=member_id)

        fs = FileSystemStorage()

        if members.count() > 0:
            member = members[0]
            if member.filename != '':
                fs.delete(member.filename)
            elif member.photo_url != '' and '/static/images/ic_profile.png' not in member.photo_url:
                fname = member.photo_url.replace(settings.URL + '/media/', '')
                fs.delete(fname)
            member.delete()

            contacts = Contact.objects.filter(member_id=member_id)
            for contact in contacts:
                contact.delete()

            gms = GroupMember.objects.filter(member_id=member_id)
            for gm in gms:
                groups = Group.objects.filter(id=gm.group_id)
                if groups.count() > 0:
                    group = groups[0]
                    if int(group.member_count) > 0:
                        group.member_count = int(group.member_count) - 1
                gm.delete()

            gcs = GroupConnect.objects.filter(member_id=member_id)
            for gc in gcs:
                gc.delete()

            posts = Post.objects.filter(member_id=member_id)
            for post in posts:
                pps = PostPicture.objects.filter(post_id=post.pk)
                for pp in pps:
                    if pp.picture_url != '':
                       fname = pp.picture_url.replace(settings.URL + '/media/', '')
                       fs.delete(fname)
                    pp.delete()
                post.delete()

            comments = Comment.objects.filter(member_id=member_id)
            for comment in comments:
                comment.delete()

            pls = PostLike.objects.filter(member_id=member_id)
            for pl in pls:
                pl.delete()

            notis = Notification.objects.filter(member_id=member_id)
            for noti in notis:
                noti.delete()

            notis = Notification.objects.filter(sender_id=member_id)
            for noti in notis:
                noti.delete()

            return redirect('/manager/home')
        else:
            return render(request, 'motherwise/result.html',
                          {'response': 'This member doesn\'t exist. Please refresh the site.'})



def active_members(request):
    import datetime
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    members = Member.objects.filter(Q(admin_id=adminID) & ~Q(registered_time='')).order_by('-id')
    for member in members:
        if member.registered_time != '':
            member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
    users, range, last_page = get_all_member_data(members)
    return render(request, 'motherwise/adminhome2.html', {'me':admin,'users':users, 'cohorts':cohorts, 'range': range, 'current': '1', 'last_page':str(last_page), 'title':'Active members'})


def inactive_members(request):
    import datetime
    try:
        if request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    members = Member.objects.filter(admin_id=adminID, registered_time='').order_by('-id')
    for member in members:
        if member.registered_time != '':
            member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
    users, range, last_page = get_all_member_data(members)
    return render(request, 'motherwise/adminhome2.html', {'me':admin,'users':users, 'cohorts':cohorts, 'range': range, 'current': '1', 'last_page':str(last_page), 'title':'Inactive members'})



@api_view(['GET', 'POST'])
def message_to_selected_members(request):

    import datetime

    if request.method == 'POST':

        ids = request.POST.getlist('users2[]')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        c = Cohort.objects.filter(admin_id=admin.pk).first()
        cohorts = []
        if c is not None:
            if c.cohorts != '': cohorts = c.cohorts.split(',')

        try:
            option = request.POST.get('option','')
            if option == 'private_chat':
                memberList = []
                memberIdList = []
                for member_id in ids:
                    members = Member.objects.filter(id=int(member_id))
                    if members.count() > 0:
                        member = members[0]
                        memberList.append(member)
                        memberIdList.append(member.pk)

                contacts = update_admin_contact(admin, "")

                if len(memberList) > 0:
                    request.session['selected_option'] = option
                    request.session['selected_member_list'] = memberIdList
                    return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'friend':memberList[0], 'contacts':contacts, 'cohorts':cohorts})
                else:
                    return redirect('/manager/home')

            elif option == 'group_chat':
                memberList = []
                memberList2 = []
                memberIdList = []
                for member_id in ids:
                    members = Member.objects.filter(id=int(member_id))
                    if members.count() > 0:
                        member = members[0]
                        memberList.append(member)
                        memberIdList.append(member.pk)
                if len(memberList) > 0:
                    request.session['selected_option'] = option
                    request.session['selected_member_list'] = memberIdList
                    groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
                    for group in groups:
                        group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                        group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                    latestGroupMemberList = []
                    latest_group = None
                    if groups.count() > 0:
                        latest_group = groups[0]
                        gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                        for gMember in gMembers:
                            members = Member.objects.filter(id=gMember.member_id)
                            if members.count() > 0:
                                latestGroupMemberList.append(members[0])
                        for memb in memberList:
                            gms = GroupMember.objects.filter(group_id=latest_group.pk, member_id=memb.pk)
                            if gms.count() == 0: memberList2.append(memb)
                    else:
                        memberList2 = memberList
                    gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
                    recents = []
                    for gc in gcs:
                        gs = Group.objects.filter(id=gc.group_id)
                        if gs.count() > 0:
                            group = gs[0]
                            group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                            group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                            recents.append(group)

                    return render(request, 'motherwise/groups.html', {'members':memberList2, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
                else:
                    return redirect('/manager/home')


        except KeyError:
            print('no such key')

        message = request.POST.get('message', '')

        for member_id in ids:
            members = Member.objects.filter(id=int(member_id))
            if members.count() > 0:
                member = members[0]

                notification = Notification()
                notification.member_id = member.pk
                notification.sender_id = admin.pk
                notification.message = message
                notification.notified_time = str(int(round(time.time() * 1000)))
                notification.save()

                rcv = Received()
                rcv.member_id = member.pk
                rcv.sender_id = admin.pk
                rcv.noti_id = notification.pk
                rcv.save()

                snt = Sent()
                snt.member_id = member.pk
                snt.sender_id = admin.pk
                snt.noti_id = notification.pk
                snt.save()

                title = 'MotherWise Community: The Nest'
                subject = 'You\'ve received a message in the Nest.(has recibido un mensaje en el Nest.)'
                msg = 'Dear ' + member.name + ', You\'ve received a message in the Nest. The message is as following:<br><br>'
                msg = msg + message
                msg = msg + '<br><br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>'

                title2 = 'Comunidad MotherWise: el Nest'
                msg2 = member.name + ', has recibido un mensaje en el Nest. el mensaje es el siguiente:<br><br>'
                msg2 = msg2 + message
                msg2 = msg2 + '<br><br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>unirse al sitio web</a>'

                from_email = admin.email
                to_emails = []
                to_emails.append(member.email)
                send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
                msg = msg + '\n\n'
                msg = msg + member.name + ', has recibido un mensaje en el Nest.\nel mensaje es el siguiente:\n' + message

                ##########################################################################################################################################################################

                db = firebase.database()
                data = {
                    "msg": message,
                    "date":str(int(round(time.time() * 1000))),
                    "sender_id": str(admin.pk),
                    "sender_name": admin.name,
                    "sender_email": admin.email,
                    "sender_photo": admin.photo_url,
                    "role": "admin",
                    "type": "message",
                    "id": str(notification.pk),
                    "mes_id": str(notification.pk)
                }

                db.child("notify").child(str(member.pk)).push(data)
                db.child("notify2").child(str(member.pk)).push(data)

                sendFCMPushNotification(member.pk, admin.pk, message)

                #################################################################################################################################################################################

                if member.playerID != '':
                    playerIDList = []
                    playerIDList.append(member.playerID)
                    url = '/mothers/notifications?noti_id=' + str(notification.pk)
                    send_push(playerIDList, msg, url)

        members = Member.objects.filter(admin_id=adminID).order_by('-id')
        users, range, last_page = get_all_member_data(members)

        return render(request, 'motherwise/adminhome2.html', {'me':admin,'users':users, 'cohorts':cohorts, 'range': range, 'last_page':str(last_page), 'current': '1', 'notify':'message_sent'})

    else:
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        c = Cohort.objects.filter(admin_id=admin.pk).first()
        cohorts = []
        if c is not None:
            if c.cohorts != '': cohorts = c.cohorts.split(',')

        memberIdList = []
        try:
            memberIdList = request.session['selected_member_list']
        except KeyError:
            print('No key')

        memberList = []
        for member_id in memberIdList:
            members = Member.objects.filter(id=member_id)
            if members.count() > 0:
                member = members[0]
                memberList.append(member)
        selectedOption = request.session['selected_option']

        contacts = update_admin_contact(admin, "")

        if len(memberList) == 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'The members don\'t exist.'})

        if len(memberList) > 0:
            if selectedOption == 'private_chat':
                return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'cohorts':cohorts, 'friend':memberList[0], 'contacts':contacts})
            elif selectedOption == 'group_chat':
                groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
                for group in groups:
                    group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                latestGroupMemberList = []
                latest_group = None
                if groups.count() > 0:
                    latest_group = groups[0]
                    gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                    for gMember in gMembers:
                        members = Member.objects.filter(id=gMember.member_id)
                        if members.count() > 0:
                            latestGroupMemberList.append(members[0])
                gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
                recents = []
                for gc in gcs:
                    gs = Group.objects.filter(id=gc.group_id)
                    if gs.count() > 0:
                        group = gs[0]
                        group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                        group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                        recents.append(group)
                return render(request, 'motherwise/groups.html', {'members':memberList, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
            else:
                return redirect('/manager/home')
        else:
            return redirect('/manager/home')



def to_page(request):
    index = request.GET['index']
    page = request.GET['page']

    import datetime

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    i = 0
    if page == 'all_members':
        if int(index) == 1:
            return redirect('/manager/home')
        userList = []
        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        for user in users:
            if user.registered_time != '':
                user.registered_time = datetime.datetime.fromtimestamp(float(int(user.registered_time)/1000)).strftime("%b %d, %Y")
        userList = users[(int(index) - 1) * 25 : int(index) * 25 - 1]
        first_page = 1
        last_page = int(users.count() / 25)
        if users.count() % 25 > 0: last_page += 1
        s = int(index) - 3
        if s < 1: s = 1
        e = int(index) + 3
        if e > last_page: e = last_page
        return render(request, 'motherwise/adminhome2.html', {'users':userList, 'cohorts':cohorts, 'range':range(s, e + 1), 'current':index, 'last_page':str(last_page)})

    elif page == 'active_members':
        if int(index) == 1:
            return redirect('/manager/active_members')
        userList = []
        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        for user in users:
            if user.registered_time != '':
                user.registered_time = datetime.datetime.fromtimestamp(float(int(user.registered_time)/1000)).strftime("%b %d, %Y")
        userList = users[(int(index) - 1) * 25 : int(index) * 25 - 1]
        first_page = 1
        last_page = int(users.count() / 25)
        if users.count() % 25 > 0: last_page += 1
        s = int(index) - 3
        if s < 1: s = 1
        e = int(index) + 3
        if e > last_page: e = last_page
        return render(request, 'motherwise/adminhome2.html', {'users':userList, 'cohorts':cohorts, 'range':range(s, e + 1), 'current':index, 'last_page':str(last_page)})

    elif page == 'inactive_members':
        if int(index) == 1:
            return redirect('/manager/inactive_members')
        userList = []
        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        for user in users:
            if user.registered_time != '':
                user.registered_time = datetime.datetime.fromtimestamp(float(int(user.registered_time)/1000)).strftime("%b %d, %Y")
        userList = users[(int(index) - 1) * 25 : int(index) * 25 - 1]
        first_page = 1
        last_page = int(users.count() / 25)
        if users.count() % 25 > 0: last_page += 1
        s = int(index) - 3
        if s < 1: s = 1
        e = int(index) + 3
        if e > last_page: e = last_page
        return render(request, 'motherwise/adminhome2.html', {'users':userList, 'cohorts':cohorts, 'range':range(s, e + 1), 'current':index, 'last_page':str(last_page)})


def to_previous(request):
    index = request.GET['index']
    page = request.GET['page']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    if page == 'all_members':
        if int(index) == 1:
            return redirect('manager/home')
    elif page == 'active_members':
        if int(index) == 1:
            return redirect('/manager/active_members')
    elif page == 'inactive_members':
        if int(index) == 1:
            return redirect('/manager/inactive_members')

    index = int(index) - 1
    return redirect('/to_page?index=' + str(index) + '&page=' + page)


def to_next(request):
    index = request.GET['index']
    page = request.GET['page']

    try:
        if request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    count = 0
    if page == 'all_members':
        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        count = users.count()

    elif page == 'active_members':
        users = Member.objects.filter(Q(admin_id=admin.pk) & ~Q(registered_time='')).order_by('-id')
        count = users.count()

    elif page == 'inactive_members':
        users = Member.objects.filter(admin_id=admin.pk, registered_time='').order_by('-id')
        count = users.count()

    r = int(count / 25)
    m = count % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    if int(index) < r - 1:
        index = int(index) + 1
    return redirect('/to_page?index=' + str(index) + '&page=' + page)



@api_view(['GET', 'POST'])
def do_cohort(request):

    import datetime

    if request.method == 'POST':
        try:
            cohort = request.POST.get('cohort','')
            option = request.POST.get('option','')
        except AssertionError:
            return redirect('/manager/home')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        if cohort == '':
            return render(request, 'motherwise/result.html',
                          {'response': 'Please choose a cohort.'})

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        c = Cohort.objects.filter(admin_id=admin.pk).first()
        cohorts = []
        if c is not None:
            if c.cohorts != '': cohorts = c.cohorts.split(',')

        members = Member.objects.filter(Q(admin_id=adminID) & Q(cohort=cohort) & ~Q(registered_time='')).order_by('-id')
        for member in members:
            if member.registered_time != '':
                member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")

        memberList = []
        memberIdList = []

        memberList = members[:25]
        for m in memberList: memberIdList.append(m.pk)
        first_page = 1
        last_page = int(members.count() / 25)
        if members.count() % 25 > 0: last_page += 1
        s = 1
        e = 7
        if e > last_page: e = last_page

        if len(memberList) == 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'The cohort\'s members don\'t exist.'})

        request.session['selected_member_list'] = memberIdList
        request.session['selected_option'] = option
        request.session['last_page'] = str(last_page)

        if option == 'members':
            return render(request, 'motherwise/adminhome2.html', {'me':admin,'users':memberList, 'cohorts':cohorts, 'range': range(s, e + 1), 'current': 1, 'cohort': cohort, 'last_page':str(last_page)})
        elif option == 'video':
            return redirect('/manager/open_conference?group_id=0&cohort=' + cohort)
        elif option == 'group_chat':
            groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
            for group in groups:
                group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
            latestGroupMemberList = []
            latest_group = None
            if groups.count() > 0:
                latest_group = groups[0]
                gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        latestGroupMemberList.append(members[0])
            gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
            recents = []
            for gc in gcs:
                gs = Group.objects.filter(id=gc.group_id)
                if gs.count() > 0:
                    group = gs[0]
                    group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                    recents.append(group)
            return render(request, 'motherwise/groups.html', {'members':memberList, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
        elif option == 'private_chat':
            contacts = update_admin_contact(admin, "")
            return render(request, 'motherwise/chat.html', {'members':memberList, 'cohorts':cohorts, 'me': admin, 'friend':memberList[0], 'contacts':contacts})

    else:
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        c = Cohort.objects.filter(admin_id=admin.pk).first()
        cohorts = []
        if c is not None:
            if c.cohorts != '': cohorts = c.cohorts.split(',')

        memberIdList = []
        try:
            memberIdList = request.session['selected_member_list']
        except KeyError:
            print('No key')

        selectedOption = request.session['selected_option']

        memberList = []
        for member_id in memberIdList:
            members = Member.objects.filter(id=member_id)
            if members.count() > 0:
                member = members[0]
                if member.registered_time != '':
                    member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
                memberList.append(member)

        contacts = update_admin_contact(admin, "")

        if len(memberList) > 0:
            if selectedOption == 'private_chat':
                return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'cohorts':cohorts, 'friend':memberList[0], 'contacts':contacts})
            elif selectedOption == 'group_chat':
                groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
                for group in groups:
                    group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                latestGroupMemberList = []
                latest_group = None
                if groups.count() > 0:
                    latest_group = groups[0]
                    gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                    for gMember in gMembers:
                        members = Member.objects.filter(id=gMember.member_id)
                        if members.count() > 0:
                            latestGroupMemberList.append(members[0])
                gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
                recents = []
                for gc in gcs:
                    gs = Group.objects.filter(id=gc.group_id)
                    if gs.count() > 0:
                        group = gs[0]
                        group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                        group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                        recents.append(group)
                return render(request, 'motherwise/groups.html', {'members':memberList, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
            elif selectedOption == 'members':
                last_page = 7
                try: last_page = int(request.session['last_page'])
                except: pass
                s = 1
                e = 7
                if e > last_page: e = last_page
                cohort = memberList[0].cohort
                return render(request, 'motherwise/adminhome2.html', {'me':admin,'users':memberList, 'cohorts':cohorts, 'range': range(s, e + 1), 'current': '1', 'last_page':str(last_page), 'cohort': cohort})
            elif selectedOption == 'video':
                cohort = memberList[0].cohort
                return redirect('/manager/open_conference?group_id=0&cohort=' + cohort)
            else:
                return redirect('/manager/home')
        else:
            return redirect('/manager/home')




@api_view(['GET', 'POST'])
def search_members(request):
    if request.method == 'POST':
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        c = Cohort.objects.filter(admin_id=admin.pk).first()
        cohorts = []
        if c is not None:
            if c.cohorts != '': cohorts = c.cohorts.split(',')

        search_id = request.POST.get('q', None)

        memberList = []

        members = Member.objects.filter(admin_id=adminID).order_by('-id')
        memberList = get_filtered_members_data(members, search_id)
        users, range, last_page = get_member_data(memberList)
        return render(request, 'motherwise/adminhome2.html', {'me':admin,'users':users, 'cohorts':cohorts, 'range': range, 'current': '1', 'last_page':str(last_page), 'title':'Searched by ' + search_id})



def get_filtered_members_data(members, keyword):
    import datetime
    memberList = []
    for member in members:
        if keyword.lower() in member.name.lower():
            memberList.append(member)
        elif keyword.lower() in member.name.lower():
            memberList.append(member)
        elif keyword.lower() in member.email.lower():
            memberList.append(member)
        elif keyword.lower() in member.phone_number.lower():
            memberList.append(member)
        elif keyword.lower() in member.cohort.lower():
            memberList.append(member)
        elif keyword.lower() in member.address.lower():
            memberList.append(member)
        else:
            if member.registered_time != '':
                date_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y %H:%M")
                if keyword.lower() in date_time.lower():
                    memberList.append(member)

    return memberList


def get_member_data(members):
    import datetime
    i = 0
    memberList = []
    for member in members:
        if member.registered_time != '':
            member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")

    memberList = members[:25]
    first_page = 1
    last_page = int(len(members) / 25)
    if len(members) % 25 > 0: last_page += 1
    s = 1
    e = 7
    if e > last_page: e = last_page

    return memberList, range(s, e + 1), last_page



def admin_account(request):
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']

    admins = Member.objects.filter(id=adminID)
    if admins.count() > 0:
        admin = admins[0]
    return  render(request, 'motherwise/account.html', {'admin':admin})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def edit_admin_account(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        oldpassword = request.POST.get('oldpassword', '')
        newpassword = request.POST.get('newpassword', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']

        admin = Member.objects.get(id=adminID)
        if email == admin.email and oldpassword == admin.password:
            admin.password = newpassword

            admin.save()

        elif email == admin.email and oldpassword != admin.password:
            return render(request, 'motherwise/result.html',
                          {'response': 'Your old password is incorrect. Please enter your correct password.'})

        else:
            return render(request, 'motherwise/result.html',
                          {'response': 'Your email or password is incorrect. Please enter your correct information.'})

        return  render(request, 'motherwise/account.html', {'admin':admin, 'note':'account_updated'})


def torequestpwd(request):
    return  render(request, 'motherwise/forgot_password.html')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def send_mail_forgotpassword(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')

        members = Member.objects.filter(email=email, cohort='admin')
        if members.count() == 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'This email doesn\'t exist for admin. Please try another one.'})

        message = 'You are allowed to reset your password from your request.<br>For it, please click this link to reset your password.<br><br><a href=\'' + 'https://www.vacay.company/resetpassword?email=' + email
        message = message + '\' target=\'_blank\'>' + 'Link to reset password' + '</a>'

        html =  """\
                    <html>
                        <head></head>
                        <body>
                            <a href="#"><img src="https://www.vacay.company/static/images/logo.png" style="width:120px;height:120px; margin-left:25px;"/></a>
                            <h2 style="color:#02839a;">MotherWise Administrator's Security Update Information</h2>
                            <div style="font-size:14px; white-space: pre-line; word-wrap: break-word;">
                                {mes}
                            </div>
                        </body>
                    </html>
                """
        html = html.format(mes=message)

        fromEmail = 'cayleywetzig@gmail.com'
        toEmailList = []
        toEmailList.append(email)
        msg = EmailMultiAlternatives('We allowed you to reset your password', '', fromEmail, toEmailList)
        msg.attach_alternative(html, "text/html")
        msg.send(fail_silently=False)

        return render(request, 'motherwise/result.html',
                          {'response': 'We sent a message to your email. Please check and reset your password.'})


def resetpassword(request):
    email = request.GET['email']
    return render(request, 'motherwise/resetpwd.html', {'email':email})




@api_view(['GET', 'POST'])
def admin_rstpwd(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')

        members = Member.objects.filter(email=email)
        if members.count() == 0:
            return redirect('/manager/signuppage/')

        member = members[0]
        member.password = password
        member.save()

        return render(request, 'motherwise/admin.html', {'notify':'password changed'})





@api_view(['GET', 'POST'])
def send_cohort_message(request):

    import datetime

    if request.method == 'POST':

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        c = Cohort.objects.filter(admin_id=admin.pk).first()
        cohorts = []
        if c is not None:
            if c.cohorts != '': cohorts = c.cohorts.split(',')

        cohort = request.POST.get('cohort', '')
        message = request.POST.get('message', '')

        members = Member.objects.filter(cohort=cohort)

        for member in members:

            notification = Notification()
            notification.member_id = member.pk
            notification.sender_id = admin.pk
            notification.message = message
            notification.notified_time = str(int(round(time.time() * 1000)))
            notification.save()

            rcv = Received()
            rcv.member_id = member.pk
            rcv.sender_id = admin.pk
            rcv.noti_id = notification.pk
            rcv.save()

            snt = Sent()
            snt.member_id = member.pk
            snt.sender_id = admin.pk
            snt.noti_id = notification.pk
            snt.save()

            title = 'You\'ve received a message in the Nest'
            subject = 'From MotherWise Community: The Nest (de la comunidad MotherWise: el Nest)'
            msg = 'Dear ' + member.name + ', You\'ve received a message in the Nest. The message is as following:<br><br>'
            msg = msg + message
            msg = msg + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>MotherWise Community'

            title2 = 'has recibido un mensaje en el Nest.'
            msg2 = member.name + ', has recibido un mensaje en el Nest. el mensaje es el siguiente:<br><br>'
            msg2 = msg2 + message
            msg2 = msg2 + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>unirse al sitio web</a>' + '<br><br>comunidad MotherWise: el Nest'

            from_email = admin.email
            to_emails = []
            to_emails.append(member.email)
            send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

            msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
            msg = msg + '\n\n'
            msg = member.name + ', has recibido un mensaje en el Nest.\nel mensaje es el siguiente:\n' + message

            ##########################################################################################################################################################################

            db = firebase.database()
            data = {
                "msg": msg,
                "date":str(int(round(time.time() * 1000))),
                "sender_id": str(admin.pk),
                "sender_name": admin.name,
                "sender_email": admin.email,
                "sender_photo": admin.photo_url,
                "role": "admin",
                "type": "message",
                "id": str(notification.pk),
                "mes_id": str(notification.pk)
            }

            db.child("notify").child(str(member.pk)).push(data)
            db.child("notify2").child(str(member.pk)).push(data)

            sendFCMPushNotification(member.pk, admin.pk, msg)

            #################################################################################################################################################################################

            if member.playerID != '':
                playerIDList = []
                playerIDList.append(member.playerID)
                url = '/mothers/notifications?noti_id=' + str(notification.pk)
                send_push(playerIDList, msg, url)

        members = Member.objects.filter(admin_id=adminID).order_by('-id')
        for member in members:
            if member.registered_time != '':
                member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
        users, range, last_page = get_all_member_data(members)

        return render(request, 'motherwise/adminhome2.html', {'me':admin,'users':users, 'cohorts':cohorts, 'range': range, 'last_page':str(last_page), 'current': '1', 'notify':'message_sent'})


def admin_switch_chat(request):

    member_id = request.GET['member_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    members = Member.objects.filter(id=member_id)
    if members.count() == 0:
        return redirect('/manager/home')

    selected_member = members[0]

    memberIdList = []
    try:
        memberIdList = request.session['selected_member_list']
    except KeyError:
        print('No key')

    selectedOption = request.session['selected_option']

    if len(memberIdList) == 0:
        return redirect('/manager/home')

    memberList = []
    for member_id in memberIdList:
        members = Member.objects.filter(id=member_id)
        if members.count() > 0:
            member = members[0]
            memberList.append(member)

    for member in memberList:
        if selected_member.pk == member.pk:
            index = memberList.index(member)
            del memberList[index]
            memberList.insert(0, selected_member)

    if selected_member not in memberList:
        memberList.insert(0,selected_member)
        memberIdList.insert(0, selected_member.pk)

    contacts = update_admin_contact(admin, "")

    if len(memberList) == 0:
        return render(request, 'motherwise/result.html', {'response': 'The member doesn\'t exist.'})

    if selectedOption == 'private_chat':
        return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'cohorts':cohorts, 'friend':memberList[0], 'contacts':contacts})
    else:
        return redirect('/manager/home')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def admin_to_chat(request):
    if request.method == 'POST':

        email = request.POST.get('member_email', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        c = Cohort.objects.filter(admin_id=admin.pk).first()
        cohorts = []
        if c is not None:
            if c.cohorts != '': cohorts = c.cohorts.split(',')

        members = Member.objects.filter(email=email)
        if members.count() == 0:
            return redirect('/manager/home')

        member = members[0]
        contacts = update_admin_contact(admin, email)

        memberList = []
        memberList.insert(0, member)

        return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'cohorts':cohorts, 'friend':memberList[0], 'contacts':contacts})

    else:
        return redirect('/manager/home')



def update_admin_contact(admin, member_email):
    if member_email != '':
        contacts = Contact.objects.filter(member_id=admin.pk, contact_email=member_email)
        if contacts.count() == 0:
            contact = Contact()
            contact.member_id = admin.pk
            contact.contact_email = member_email
            contact.contacted_time = str(int(round(time.time() * 1000)))
            contact.save()
        else:
            contact = contacts[0]
            contacts = Contact.objects.filter(member_id=admin.pk)
            recent_contact = contacts[contacts.count() - 1]
            if contact.pk < recent_contact.pk:
                contact.delete()
                contact = Contact()
                contact.member_id = admin.pk
                contact.contact_email = member_email
                contact.contacted_time = str(int(round(time.time() * 1000)))
                contact.save()

    contacts = Contact.objects.filter(member_id=admin.pk).order_by('-id')
    contactList = []
    for contact in contacts:
        members = Member.objects.filter(email=contact.contact_email)
        if members.count() > 0:
            member = members[0]
            contactList.append(member)

    return contactList



def admin_switch_to_cohort(request):

    cohort = request.GET['cohort']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    members = Member.objects.filter(admin_id=adminID).order_by('-id')
    memberList = []
    memberIdList = []
    for member in members:
        if member.cohort.lower() == cohort.lower():
            memberList.append(member)
            memberIdList.append(member.pk)

    if len(memberList) == 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'The cohort\'s members don\'t exist.'})

    request.session['selected_member_list'] = memberIdList

    contacts = update_admin_contact(admin, "")

    return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'cohorts':cohorts, 'friend':memberList[0], 'contacts':contacts})




@api_view(['GET', 'POST'])
def create_group(request):

    import datetime

    if request.method == 'POST':

        name = request.POST.get('name', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        groups = Group.objects.filter(member_id=admin.pk, name=name)
        if groups.count() > 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'The same name already exists.'})

        memberIdList = []
        try:
            memberIdList = request.session['selected_member_list']
        except KeyError:
            print('No key')

        memberList = []
        for member_id in memberIdList:
            members = Member.objects.filter(id=member_id)
            if members.count() > 0:
                member = members[0]
                memberList.append(member)

        group = Group()
        group.member_id = admin.pk
        group.name = name
        group.code = get_group_code(name)
        group.color = get_group_color()
        group.member_count = '0'
        group.created_time = str(int(round(time.time() * 1000)))
        group.last_connected_time = str(int(round(time.time() * 1000)))
        group.save()

        groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
        for group in groups:
            group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
            group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
        latestGroupMemberList = []
        latest_group = None
        if groups.count() > 0:
            latest_group = groups[0]
            gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
            for gMember in gMembers:
                members = Member.objects.filter(id=gMember.member_id)
                if members.count() > 0:
                    latestGroupMemberList.append(members[0])
        gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
        recents = []
        for gc in gcs:
            gs = Group.objects.filter(id=gc.group_id)
            if gs.count() > 0:
                g = gs[0]
                g.created_time = datetime.datetime.fromtimestamp(float(int(g.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                g.last_connected_time = datetime.datetime.fromtimestamp(float(int(g.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                recents.append(g)
        return render(request, 'motherwise/groups.html', {'members':memberList, 'group':group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})

    else:
        return redirect('/manager/home')


def get_group_code(name):
    name = name.split(' ')
    if len(name) > 1:
        name = name[0][0:1] + name[1][0:1]
    else: name = name[0][0:1]
    name = name.upper()
    return name

def get_group_color():
    from random import randint
    color = '#{:06x}'.format(randint(0, 256**3))
    return color



@api_view(['GET', 'POST'])
def message_to_group(request):

    import datetime

    if request.method == 'POST':

        group_id = request.POST.get('group_id', '1')
        message = request.POST.get('message', '')
        option = request.POST.get('option', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        groups = Group.objects.filter(id=group_id)
        if groups.count() > 0:
            group = groups[0]

            if option == 'invite':

                memberIdList = []
                try:
                    memberIdList = request.session['selected_member_list']
                except KeyError:
                    print('No key')

                memberList = []
                for member_id in memberIdList:
                    members = Member.objects.filter(id=member_id)
                    if members.count() > 0:
                        member = members[0]
                        gMembers = GroupMember.objects.filter(group_id=group.pk, member_id=member.pk)
                        if gMembers.count() == 0:
                            gMember = GroupMember()
                            gMember.group_id = group.pk
                            gMember.member_id = member.pk
                            gMember.invited_time = str(int(round(time.time() * 1000)))
                            gMember.save()

                            notification = Notification()
                            notification.member_id = member.pk
                            notification.sender_id = admin.pk
                            notification.notified_time = str(int(round(time.time() * 1000)))
                            notification.save()

                            rcv = Received()
                            rcv.member_id = member.pk
                            rcv.sender_id = admin.pk
                            rcv.noti_id = notification.pk
                            rcv.save()

                            snt = Sent()
                            snt.member_id = member.pk
                            snt.sender_id = admin.pk
                            snt.noti_id = notification.pk
                            snt.save()

                            title = 'You\'ve received an invitation from MotherWise Community: The Nest'
                            subject = 'From MotherWise Community: The Nest'
                            msg = 'Dear ' + member.name + ',<br><br>You\'ve received an invitation to: ' + group.name + ' from MotherWise Community.<br><br>Community name: ' + group.name + '<br><br>'
                            # msg = msg + 'So you can see that community in your account and connect it to became a member of the community, attending all the events from it.<br>'
                            msg = msg + 'There’s nothing you have to do- you’re already included. The next time you login, just click on the \"Communities\" icon and you’ll see this option.<br>'
                            msg = msg + 'From there, you’ll see the members, see which videos have been posted, engage in a group chat and send private messages.<br>'
                            msg = msg + 'These communities are ways you can share similar interests and get to know your Nest Family. Have fun!<br>'
                            msg = msg + '<a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>MotherWise Team'

                            title2 = 'Recibiste una invitación de MotherWise Community: The Nest'
                            msg2 = 'Querida ' + member.name + ',<br><br>Recibió una invitación para una comunidad ' + group.name + ' de la comunidad MotherWise.<br><br>Nombre de la comunidad: ' + group.name + '<br><br>'
                            # msg2 = msg2 + 'No hay nada que tenga que hacer, ya está incluida. La próxima vez que inicie sesión, simplemente haga clic en el icono de la comunidad y verá esta opción.<br>'
                            msg2 = msg2 + 'No hay nada que tengas que hacer, ya estás incluido. La próxima vez que inicie sesión, simplemente haga clic en el icono \"Comunidades\" y verá esta opción.<br>'
                            msg2 = msg2 + 'A partir de ahí, verá los miembros, verá qué videos se han publicado, participará en un chat grupal y enviará mensajes privados.<br>'
                            msg2 = msg2 + 'Estas comunidades son formas en que puede compartir intereses similares y conocer a su familia Nest. ¡Que te diviertas!<br>'
                            msg2 = msg2 + '<a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Unirse al sitio web</a>' + '<br><br>MotherWise Team'

                            from_email = admin.email
                            to_emails = []
                            to_emails.append(member.email)
                            # send_mail_message(from_email, to_emails, title, subject, msg)
                            send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                            msg = 'Dear ' + member.name + ',\n\nYou\'ve received an invitation to: ' + group.name + ' from MotherWise Community.\n\nCommunity name: ' + group.name + '\n\n'
                            # msg = msg + 'So you can see that community in your account and connect it to became a member of the community, attending all the events from it.\n'
                            msg = msg + 'There’s nothing you have to do- you’re already included. The next time you login, just click on the \"Communities\" icon and you’ll see this option.\n'
                            msg = msg + 'From there, you’ll see the members, see which videos have been posted, engage in a group chat and send private messages.\n'
                            msg = msg + 'These communities are ways you can share similar interests and get to know your Nest Family. Have fun!\n'
                            msg = msg + 'Click on this link to join: ' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\n\nMotherWise Team'

                            msg2 = member.name + ',\n\nha recibido una invitación para:' + group.name + ' de la comunidad MotherWise.\n\nNombre de la comunidad: ' + group.name + '\n\n'
                            # msg2 = msg2 + 'So you can see that community in your account and connect it to became a member of the community, attending all the events from it.\n'
                            msg2 = msg2 + 'No tienes que hacer nada, ya estás incluido. La próxima vez que inicie sesión, simplemente haga clic en el icono "Comunidades" y verá esta opción.\n'
                            msg2 = msg2 + 'Desde allí, verá a los miembros, verá qué videos se han publicado, participará en un chat grupal y enviará mensajes privados.\n'
                            msg2 = msg2 + 'Estas comunidades son formas en las que puede compartir intereses similares y conocer a su familia Nest. ¡Divertirse!\n'
                            msg2 = msg2 + 'Haga clic en este enlace para unirse: ' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\n\nEquipo MotherWise'

                            notification.message = msg + '\n\n' + msg2
                            notification.save()

                            ##########################################################################################################################################################################

                            msg = msg + '\n\n' + msg2

                            db = firebase.database()
                            data = {
                                "msg": msg,
                                "date":str(int(round(time.time() * 1000))),
                                "sender_id": str(admin.pk),
                                "sender_name": admin.name,
                                "sender_email": admin.email,
                                "sender_photo": admin.photo_url,
                                "role": "admin",
                                "type": "group_invite",
                                "id": str(group.pk),
                                "mes_id": str(notification.pk)
                            }

                            db.child("notify").child(str(member.pk)).push(data)
                            db.child("notify2").child(str(member.pk)).push(data)

                            sendFCMPushNotification(member.pk, admin.pk, msg)

                            #################################################################################################################################################################################

                            if member.playerID != '':
                                playerIDList = []
                                playerIDList.append(member.playerID)
                                url = '/mothers/notifications?noti_id=' + str(notification.pk)
                                send_push(playerIDList, msg, url)

                        gms = GroupMember.objects.filter(group_id=group.pk, member_id=member.pk)
                        if gms.count() == 0: memberList.append(member)

                gMembers = GroupMember.objects.filter(group_id=group.pk).order_by('-id')
                group.member_count = str(gMembers.count())
                group.last_connected_time = str(int(round(time.time() * 1000)))
                group.save()

                update_recent_group(admin, group)

                latestGroupMemberList = []
                groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
                for gr in groups:
                    gr.created_time = datetime.datetime.fromtimestamp(float(int(gr.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    gr.last_connected_time = datetime.datetime.fromtimestamp(float(int(gr.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        latestGroupMemberList.append(members[0])

                gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
                recents = []
                for gc in gcs:
                    gs = Group.objects.filter(id=gc.group_id)
                    if gs.count() > 0:
                        g = gs[0]
                        g.created_time = datetime.datetime.fromtimestamp(float(int(g.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                        g.last_connected_time = datetime.datetime.fromtimestamp(float(int(g.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                        recents.append(g)

                return render(request, 'motherwise/groups.html', {'members':[], 'group':group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})

            elif option == 'group':
                gMembers = GroupMember.objects.filter(group_id=group.pk).order_by('-id')
                for gm in gMembers:
                    members = Member.objects.filter(id=gm.member_id)
                    if members.count() > 0:
                        member = members[0]

                        notification = Notification()
                        notification.member_id = member.pk
                        notification.sender_id = admin.pk
                        notification.message = message
                        notification.notified_time = str(int(round(time.time() * 1000)))
                        notification.save()

                        rcv = Received()
                        rcv.member_id = member.pk
                        rcv.sender_id = admin.pk
                        rcv.noti_id = notification.pk
                        rcv.save()

                        snt = Sent()
                        snt.member_id = member.pk
                        snt.sender_id = admin.pk
                        snt.noti_id = notification.pk
                        snt.save()

                        title = 'You\'ve received a message in the Nest'
                        subject = 'From MotherWise Community: The Nest (de la comunidad MotherWise.)'
                        msg = 'Dear ' + member.name + ', You\'ve received a message in the Nest. The message is as following:<br><br>'
                        msg = msg + message
                        msg = msg + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>MotherWise Team'

                        title2 = 'has recibido un mensaje en el Nest.'
                        msg2 = member.name + ', has recibido un mensaje en el Nest. el mensaje es el siguiente:<br><br>'
                        msg2 = msg2 + message
                        msg2 = msg2 + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>unirse al sitio web</a>' + '<br><br>Equipo MotherWise'

                        from_email = admin.email
                        to_emails = []
                        to_emails.append(member.email)
                        send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                        msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
                        msg2 = member.name + ', has recibido un mensaje en el Nest.\nel mensaje es el siguiente:\n' + message

                        msg = msg + '\n\n' + msg2
                        ##########################################################################################################################################################################

                        db = firebase.database()
                        data = {
                            "msg": message,
                            "date":str(int(round(time.time() * 1000))),
                            "sender_id": str(admin.pk),
                            "sender_name": admin.name,
                            "sender_email": admin.email,
                            "sender_photo": admin.photo_url,
                            "role": "admin",
                            "type": "message",
                            "id": str(notification.pk),
                            "mes_id": str(notification.pk)
                        }

                        db.child("notify").child(str(member.pk)).push(data)
                        db.child("notify2").child(str(member.pk)).push(data)

                        sendFCMPushNotification(member.pk, admin.pk, message)

                        #################################################################################################################################################################################

                        if member.playerID != '':
                            playerIDList = []
                            playerIDList.append(member.playerID)
                            url = '/mothers/notifications?noti_id=' + str(notification.pk)
                            send_push(playerIDList, msg, url)

                update_recent_group(admin, group)

                memberIdList = []
                try:
                    memberIdList = request.session['selected_member_list']
                except KeyError:
                    print('No key')

                memberList = []
                for member_id in memberIdList:
                    members = Member.objects.filter(id=member_id)
                    if members.count() > 0:
                        member = members[0]
                        gms = GroupMember.objects.filter(group_id=group.pk, member_id=member.pk)
                        if gms.count() == 0: memberList.append(member)

                latestGroupMemberList = []
                groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
                for gr in groups:
                    gr.created_time = datetime.datetime.fromtimestamp(float(int(gr.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    gr.last_connected_time = datetime.datetime.fromtimestamp(float(int(gr.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        latestGroupMemberList.append(members[0])

                gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
                recents = []
                for gc in gcs:
                    gs = Group.objects.filter(id=gc.group_id)
                    if gs.count() > 0:
                        g = gs[0]
                        g.created_time = datetime.datetime.fromtimestamp(float(int(g.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                        g.last_connected_time = datetime.datetime.fromtimestamp(float(int(g.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                        recents.append(g)

                return render(request, 'motherwise/groups.html', {'members':memberList, 'group':group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})

            else: return redirect('/manager/home')

        else:
            return redirect('/manager/home')

    else:
        return redirect('/manager/home')


def update_recent_group(admin, group):
    gConnects = GroupConnect.objects.filter(member_id=admin.pk, group_id=group.pk)
    if gConnects.count() > 0:
        gConnect = gConnects[0]
        recent = gConnects[gConnects.count() - 1]
        if gConnect.pk < recent.pk:
            gConnect.delete()
            gc = GroupConnect()
            gc.member_id = admin.pk
            gc.group_id = group.pk
            gc.last_connected_time = str(int(round(time.time() * 1000)))
            gc.save()
    else:
        gc = GroupConnect()
        gc.member_id = admin.pk
        gc.group_id = group.pk
        gc.last_connected_time = str(int(round(time.time() * 1000)))
        gc.save()



def switch_group(request):

    import datetime

    group_id = request.GET['group_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    request.session['url'] = settings.URL + request.get_full_path()

    groups = Group.objects.filter(id=group_id)
    if groups.count() == 0:
        return redirect('/manager/home')

    selected_group = groups[0]
    selected_group.created_time = datetime.datetime.fromtimestamp(float(int(selected_group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
    selected_group.last_connected_time = datetime.datetime.fromtimestamp(float(int(selected_group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")

    memberIdList = []
    try:
        memberIdList = request.session['selected_member_list']
    except KeyError:
        print('No key')

    selectedOption = request.session['selected_option']

    # if len(memberIdList) == 0:
    #     return redirect('/manager/home')

    memberList = []
    for member_id in memberIdList:
        members = Member.objects.filter(id=member_id)
        if members.count() > 0:
            member = members[0]
            gms = GroupMember.objects.filter(group_id=selected_group.pk, member_id=member.pk)
            if gms.count() == 0: memberList.append(member)

    groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
    groupList = []
    for group in groups:
        group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
        group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
        groupList.append(group)
    for group in groupList:
        if selected_group.pk == group.pk:
            index = groupList.index(group)
            del groupList[index]
            groupList.insert(0, selected_group)

    latestGroupMemberList = []
    gMembers = GroupMember.objects.filter(group_id=selected_group.pk)
    for gMember in gMembers:
        members = Member.objects.filter(id=gMember.member_id)
        if members.count() > 0:
            latestGroupMemberList.append(members[0])
    gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
    recents = []
    for gc in gcs:
        gs = Group.objects.filter(id=gc.group_id)
        if gs.count() > 0:
            group = gs[0]
            group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
            group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
            recents.append(group)
    return render(request, 'motherwise/groups.html', {'members':memberList, 'group':selected_group, 'groups': groupList, 'group_members':latestGroupMemberList, 'recents':recents})


def get_groups(request):
    import datetime
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    request.session['url'] = settings.URL + request.get_full_path()

    groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
    for group in groups:
        group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
        group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
    latestGroupMemberList = []
    latest_group = None
    if groups.count() > 0:
        latest_group = groups[0]
        gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
        for gMember in gMembers:
            members = Member.objects.filter(id=gMember.member_id)
            if members.count() > 0:
                latestGroupMemberList.append(members[0])
    gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
    recents = []
    for gc in gcs:
        gs = Group.objects.filter(id=gc.group_id)
        if gs.count() > 0:
            group = gs[0]
            group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
            group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
            recents.append(group)
    request.session['selected_option'] = 'group_chat'
    return render(request, 'motherwise/groups.html', {'members':[], 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})


def delete_group_member(request):
    member_id = request.GET['member_id']
    group_id = request.GET['group_id']
    gms = GroupMember.objects.filter(group_id=group_id, member_id=member_id)
    if gms.count() > 0:
        gms[0].delete()

    return redirect(request.session['url'])


def open_group_chat(request):
    group_id = request.GET['group_id']
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    memberList = []
    group = None
    groups = Group.objects.filter(member_id=admin.pk, id=group_id).order_by('-id')
    if groups.count() > 0:
        group = groups[0]
        gMembers = GroupMember.objects.filter(group_id=group.pk)
        for gMember in gMembers:
            members = Member.objects.filter(id=gMember.member_id)
            if members.count() > 0:
                memb = members[0]
                memb.username = '@' + memb.email[0:memb.email.find('@')]
                memberList.append(memb)

        request.session['group_id'] = group_id
        request.session['cohort'] = ''

        memberIdList = []
        for memb in memberList:
            memberIdList.append(memb.pk)
        request.session['selected_member_list'] = memberIdList

        return render(request, 'motherwise/group_chat.html', {'me':admin, 'members':memberList, 'cohorts':cohorts, 'group':group})
    else:
        return redirect('/manager/home')


def group_cohort_chat(request):
    cohort = request.GET['cohort']
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    members = Member.objects.filter(admin_id=admin.pk, cohort=cohort).order_by('-id')
    for member in members:
        member.username = '@' + member.email[0:member.email.find('@')]

    request.session['cohort'] = cohort
    request.session['group_id'] = ''

    memberIdList = []
    memberList = []
    for memb in members:
        if memb.registered_time != '':
            memberList.append(memb)
            memberIdList.append(memb.pk)
    request.session['selected_member_list'] = memberIdList

    return render(request, 'motherwise/group_chat.html', {'me':admin, 'members':memberList, 'cohorts':cohorts, 'cohort':cohort})



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def group_chat_message(request):
    if request.method == 'POST':

        group_id = request.POST.get('group_id', '')
        message = request.POST.get('message', '')
        option = request.POST.get('option', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        if option == 'group':
            groups = Group.objects.filter(id=group_id)
            if groups.count() > 0:
                group = groups[0]
                gMembers = GroupMember.objects.filter(group_id=group.pk)
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        member = members[0]

                        notification = Notification()
                        notification.member_id = member.pk
                        notification.sender_id = admin.pk
                        notification.message = message
                        notification.notified_time = str(int(round(time.time() * 1000)))
                        notification.save()

                        rcv = Received()
                        rcv.member_id = member.pk
                        rcv.sender_id = admin.pk
                        rcv.noti_id = notification.pk
                        rcv.save()

                        snt = Sent()
                        snt.member_id = member.pk
                        snt.sender_id = admin.pk
                        snt.noti_id = notification.pk
                        snt.save()

                        title = 'You\'ve received a message in the Nest'
                        subject = 'From MotherWise Community: The Nest (de la comunidad MotherWise: The Nest)'
                        msg = 'Dear ' + member.name + ', You\'ve received a message in the Nest. The message is as following:<br><br>'
                        msg = msg + message
                        msg = msg + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>MotherWise Team'

                        title2 = 'has recibido un mensaje en el Nest.'
                        msg2 = member.name + ', has recibido un mensaje en el Nest. el mensaje es el siguiente:<br><br>'
                        msg2 = msg2 + message
                        msg2 = msg2 + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>unirse al sitio web</a>' + '<br><br>Equipo MotherWise'

                        from_email = admin.email
                        to_emails = []
                        to_emails.append(member.email)
                        send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                        msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
                        msg2 = member.name + ', has recibido un mensaje en el Nest.\nel mensaje es el siguiente:\n' + message

                        msg = msg + msg2

                        ##########################################################################################################################################################################

                        db = firebase.database()
                        data = {
                            "msg": message,
                            "date":str(int(round(time.time() * 1000))),
                            "sender_id": str(admin.pk),
                            "sender_name": admin.name,
                            "sender_email": admin.email,
                            "sender_photo": admin.photo_url,
                            "role": "admin",
                            "type": "message",
                            "id": str(notification.pk),
                            "mes_id": str(notification.pk)
                        }

                        db.child("notify").child(str(member.pk)).push(data)
                        db.child("notify2").child(str(member.pk)).push(data)

                        sendFCMPushNotification(member.pk, admin.pk, message)

                        #################################################################################################################################################################################

                        if member.playerID != '':
                            playerIDList = []
                            playerIDList.append(member.playerID)
                            url = '/mothers/notifications?noti_id=' + str(notification.pk)
                            send_push(playerIDList, msg, url)

            else:
                return redirect('/manager/home')

        elif option == 'cohort':
            members = Member.objects.filter(cohort=group_id)
            for member in members:

                notification = Notification()
                notification.member_id = member.pk
                notification.sender_id = admin.pk
                notification.message = message
                notification.notified_time = str(int(round(time.time() * 1000)))
                notification.save()

                rcv = Received()
                rcv.member_id = member.pk
                rcv.sender_id = admin.pk
                rcv.noti_id = notification.pk
                rcv.save()

                snt = Sent()
                snt.member_id = member.pk
                snt.sender_id = admin.pk
                snt.noti_id = notification.pk
                snt.save()

                title = 'You\'ve received a message in the Nest'
                subject = 'From MotherWise Community: The Nest (de la comunidad MotherWise: The Nest)'
                msg = 'Dear ' + member.name + ', You\'ve received a message in the Nest. The message is as following:<br><br>'
                msg = msg + message
                msg = msg + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>MotherWise Team'

                title2 = 'has recibido un mensaje en el Nest.'
                msg2 = member.name + ', has recibido un mensaje en el Nest. el mensaje es el siguiente:<br><br>'
                msg2 = msg2 + message
                msg2 = msg2 + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>unirse al sitio web</a>' + '<br><br>Equipo MotherWise'

                from_email = admin.email
                to_emails = []
                to_emails.append(member.email)
                send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
                msg2 = member.name + ', has recibido un mensaje en el Nest.\nel mensaje es el siguiente:\n' + message
                msg = msg + '\n\n' + msg2
                ##########################################################################################################################################################################

                db = firebase.database()
                data = {
                    "msg": message,
                    "date":str(int(round(time.time() * 1000))),
                    "sender_id": str(admin.pk),
                    "sender_name": admin.name,
                    "sender_email": admin.email,
                    "sender_photo": admin.photo_url,
                    "role": "admin",
                    "type": "message",
                    "id": str(notification.pk),
                    "mes_id": str(notification.pk)
                }

                db.child("notify").child(str(member.pk)).push(data)
                db.child("notify2").child(str(member.pk)).push(data)

                sendFCMPushNotification(member.pk, admin.pk, message)

                #################################################################################################################################################################################

                if member.playerID != '':
                    playerIDList = []
                    playerIDList.append(member.playerID)
                    url = '/mothers/notifications?noti_id=' + str(notification.pk)
                    send_push(playerIDList, msg, url)

        cohort = request.session['cohort']
        group_id = request.session['group_id']

        if cohort != '':
            return redirect('/manager/group_cohort_chat?cohort=' + cohort)
        elif group_id is not None and int(group_id) > 0:
            return redirect('/manager/open_group_chat?group_id=' + group_id)
        else:
            return redirect('/manager/home')


def group_private_chat(request):

    email = request.GET['email']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    members = Member.objects.filter(email=email)
    if members.count() == 0:
        return redirect('/manager/home')

    member = members[0]
    contacts = update_admin_contact(admin, email)

    memberList = []
    memberList.insert(0, member)

    memberIdList = []
    memberIdList.insert(0, member.pk)

    request.session['selected_option'] = 'private_chat'
    request.session['selected_member_list'] = memberIdList

    return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'cohorts':cohorts, 'friend':memberList[0], 'contacts':contacts})



def admin_delete_contact(request):

    member_id = request.GET['member_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    members = Member.objects.filter(id=member_id)
    if members.count() > 0:
        member = members[0]
        contacts = Contact.objects.filter(member_id=admin.pk, contact_email=member.email)
        for contact in contacts:
            contact.delete()

    memberIdList = []
    try:
        memberIdList = request.session['selected_member_list']
    except KeyError:
        print('No key')

    memberList = []
    for member_id in memberIdList:
        members = Member.objects.filter(id=member_id)
        if members.count() > 0:
            member = members[0]
            memberList.append(member)

    contacts = update_admin_contact(admin, "")

    return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'cohorts':cohorts, 'friend':memberList[0], 'contacts':contacts})


def admin_delete_group(request):

    group_id = request.GET['group_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    groups = Group.objects.filter(id=group_id)
    if groups.count() > 0:
        group = groups[0]
        gms = GroupMember.objects.filter(group_id=group.pk)
        for gm in gms:
            gm.delete()

        gcs = GroupConnect.objects.filter(group_id=group.pk)
        for gc in gcs:
            gc.delete()

        group.delete()

    return redirect('/manager/groups/')



def to_posts(request):
    import datetime
    try:
        if request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
    userList = []
    for user in users:
        if user.registered_time != '':
            user.username = '@' + user.email[0:user.email.find('@')]
            userList.append(user)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    uitype = ''
    if request.user_agent.is_mobile:
        uitype = 'mobile'

    allPosts = Post.objects.filter(sch_status='').order_by('-id')
    i = 0
    itop = 1
    for post in allPosts:
        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
        i = i + 1
        pl = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk).first()
        if pl is not None: post.liked = pl.status
        else: post.liked = ''

        comments = Comment.objects.filter(post_id=post.pk, comment_id='0')
        post.comments = str(comments.count())
        likes = PostLike.objects.filter(post_id=post.pk)
        post.reactions = str(likes.count())
        post.content = emoji.emojize(post.content)

        members = Member.objects.filter(id=post.member_id)
        if members.count() > 0:
            memb = members[0]
            if int(memb.admin_id) == admin.pk or memb.pk == admin.pk:
                prevs = PostUrlPreview.objects.filter(post_id=post.pk)

                comments1 = comments[:5]
                commentlist = []
                for comment in comments1:
                    cm = Member.objects.filter(id=comment.member_id).first()
                    if cm is not None:
                        comment.comment_text = emoji.emojize(comment.comment_text)
                        commentlist.append( { 'comment':comment, 'member':cm } )

                data = {
                    'member':memb,
                    'post': post,
                    'prevs': prevs,
                    'comments': commentlist,
                    'pc-cnt': str(comments.count()),
                }

                if uitype == 'mobile':
                    if 'top' in post.status:
                        list1.insert(0,data)
                    else:
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                else:
                    if 'top' in post.status:
                        if itop == 1:
                            list1.insert(0,data)
                            itop = 2
                        elif itop == 2:
                            list2.insert(0,data)
                            itop = 3
                        elif itop == 3:
                            list3.insert(0,data)
                            itop = 4
                        elif itop == 4:
                            list4.insert(0,data)
                            itop = 1
                    else:
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)

    pst = None
    try:
        post_id = request.GET['post_id']
        posts = Post.objects.filter(id=post_id)
        if posts.count() > 0:
            pst = posts[0]
            pst.posted_time = datetime.datetime.fromtimestamp(float(int(pst.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
    except KeyError:
        print('no key')

    categories = []
    pc = PostCategory.objects.filter(admin_id=admin.pk).first()
    if pc is not None:
        if pc.categories != '': categories = pc.categories.split(',')

    return render(request, 'motherwise/post.html', {'me':admin, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'users':userList, 'pst':pst, 'categories':categories})



def my_posts(request):

    import datetime

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    uitype = ''
    if request.user_agent.is_mobile:
        uitype = 'mobile'

    users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
    userList = []
    for user in users:
        if user.registered_time != '':
            user.username = '@' + user.email[0:user.email.find('@')]
            userList.append(user)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    posts = Post.objects.filter(member_id=admin.pk).order_by('-id')

    i = 0
    itop = 1
    for post in posts:
        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")

        pl = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk).first()
        if pl is not None: post.liked = pl.status
        else: post.liked = ''

        comments = Comment.objects.filter(post_id=post.pk, comment_id='0')
        post.comments = str(comments.count())
        likes = PostLike.objects.filter(post_id=post.pk)
        post.reactions = str(likes.count())
        post.content = emoji.emojize(post.content)

        prevs = PostUrlPreview.objects.filter(post_id=post.pk)

        comments1 = comments[:5]
        commentlist = []
        for comment in comments1:
            cm = Member.objects.filter(id=comment.member_id).first()
            if cm is not None:
                comment.comment_text = emoji.emojize(comment.comment_text)
                commentlist.append( { 'comment':comment, 'member':cm } )

        i = i + 1

        data = {
            'member':admin,
            'post': post,
            'prevs': prevs,
            'comments': commentlist,
            'pc-cnt': str(comments.count()),
        }

        if uitype == 'mobile':
            if 'top' in post.status:
                list1.insert(0,data)
            else:
                if i % 4 == 1: list1.append(data)
                elif i % 4 == 2: list2.append(data)
                elif i % 4 == 3: list3.append(data)
                elif i % 4 == 0: list4.append(data)
        else:
            if 'top' in post.status:
                if itop == 1:
                    list1.insert(0,data)
                    itop = 2
                elif itop == 2:
                    list2.insert(0,data)
                    itop = 3
                elif itop == 3:
                    list3.insert(0,data)
                    itop = 4
                elif itop == 4:
                    list4.insert(0,data)
                    itop = 1
            else:
                if i % 4 == 1: list1.append(data)
                elif i % 4 == 2: list2.append(data)
                elif i % 4 == 3: list3.append(data)
                elif i % 4 == 0: list4.append(data)

    categories = []
    pc = PostCategory.objects.filter(admin_id=admin.pk).first()
    if pc is not None:
        if pc.categories != '': categories = pc.categories.split(',')

    return render(request, 'motherwise/post.html', {'me':admin, 'member':admin, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'users':userList, 'categories':categories})




def member_posts(request):
    import datetime
    try:
        if request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
    userList = []
    for user in users:
        if user.registered_time != '':
            user.username = '@' + user.email[0:user.email.find('@')]
            userList.append(user)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    uitype = ''
    if request.user_agent.is_mobile:
        uitype = 'mobile'

    user_id = request.GET['user_id']

    member = Member.objects.filter(id=user_id).first()

    uposts = Post.objects.filter(member_id=user_id, sch_status='').order_by('-id')
    i = 0
    itop = 1
    for post in uposts:
        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
        i = i + 1
        pl = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk).first()
        if pl is not None: post.liked = pl.status
        else: post.liked = ''

        comments = Comment.objects.filter(post_id=post.pk, comment_id='0')
        post.comments = str(comments.count())
        likes = PostLike.objects.filter(post_id=post.pk)
        post.reactions = str(likes.count())
        post.content = emoji.emojize(post.content)

        members = Member.objects.filter(id=post.member_id)
        if members.count() > 0:
            memb = members[0]
            if int(memb.admin_id) == admin.pk or memb.pk == admin.pk:
                prevs = PostUrlPreview.objects.filter(post_id=post.pk)

                comments1 = comments[:5]
                commentlist = []
                for comment in comments1:
                    cm = Member.objects.filter(id=comment.member_id).first()
                    if cm is not None:
                        comment.comment_text = emoji.emojize(comment.comment_text)
                        commentlist.append( { 'comment':comment, 'member':cm } )

                data = {
                    'member':memb,
                    'post': post,
                    'prevs': prevs,
                    'comments': commentlist,
                    'pc-cnt': str(comments.count()),
                }

                if uitype == 'mobile':
                    if 'top' in post.status:
                        list1.insert(0,data)
                    else:
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                else:
                    if 'top' in post.status:
                        if itop == 1:
                            list1.insert(0,data)
                            itop = 2
                        elif itop == 2:
                            list2.insert(0,data)
                            itop = 3
                        elif itop == 3:
                            list3.insert(0,data)
                            itop = 4
                        elif itop == 4:
                            list4.insert(0,data)
                            itop = 1
                    else:
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)

    pst = None
    try:
        post_id = request.GET['post_id']
        posts = Post.objects.filter(id=post_id)
        if posts.count() > 0:
            pst = posts[0]
            pst.posted_time = datetime.datetime.fromtimestamp(float(int(pst.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
    except KeyError:
        print('no key')

    categories = []
    pc = PostCategory.objects.filter(admin_id=admin.pk).first()
    if pc is not None:
        if pc.categories != '': categories = pc.categories.split(',')

    return render(request, 'motherwise/post.html', {'me':admin, 'member':member, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'users':userList, 'pst':pst, 'categories':categories})



@api_view(['GET', 'POST'])
def create_post(request):
    if request.method == 'POST':

        title = request.POST.get('title', '')
        category = request.POST.get('category', '')
        content = request.POST.get('content', '')
        ids = request.POST.getlist('users[]')
        scheduled_time = request.POST.get('scheduled_time', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        post = Post()
        post.member_id = admin.pk
        post.title = title
        post.category = category
        post.content = emoji.demojize(content)
        post.picture_url = ''
        post.comments = '0'
        post.likes = '0'
        post.loves = '0'
        post.haha = '0'
        post.wow = '0'
        post.sad = '0'
        post.angry = '0'
        post.reactions = '0'
        post.scheduled_time = scheduled_time
        post.posted_time = str(int(round(time.time() * 1000)))
        if len(ids) > 0: post.notified_members = ",".join(str(i) for i in ids)
        if scheduled_time != '': post.sch_status = 'scheduled'
        post.save()

        createPostUrlPreview(post)

        fs = FileSystemStorage()
        i = 0
        for f in request.FILES.getlist('pictures'):
            i = i + 1
            filename = fs.save(f.name, f)
            uploaded_url = fs.url(filename)
            if i == 1:
                post.picture_url = settings.URL + uploaded_url
                post.save()
            postPicture = PostPicture()
            postPicture.post_id = post.pk
            postPicture.picture_url = settings.URL + uploaded_url
            postPicture.filename = filename
            postPicture.save()


        if post.scheduled_time == '':

            for member_id in ids:
                members = Member.objects.filter(id=int(member_id))
                if members.count() > 0:
                    member = members[0]

                    title = 'MotherWise Community: The Nest'
                    subject = 'You\'ve received a post in the Nest (has recibido una publicación en el Nest)'
                    msg = 'Dear ' + member.name + ', You\'ve received a post from MotherWise Community: The Nest.<br><br>'
                    msg = msg + '<a href=\'' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk) + '\' target=\'_blank\'>View the post</a>'

                    title2 = 'comunidad MotherWise: el Nest'
                    msg2 = member.name + ', has recibido una publicación en el Nest.<br><br>'
                    msg2 = msg2 + '<a href=\'' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk) + '\' target=\'_blank\'>ver la publicación</a>'

                    from_email = admin.email
                    to_emails = []
                    to_emails.append(member.email)
                    send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                    msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\n\n'
                    msg = msg + 'Click on this link to view the post: ' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk)

                    msg2 = member.name + ', has recibido una publicación en el Nest.\n\n'
                    msg2 = msg2 + 'haga clic en este enlace para ver la publicación: ' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk)

                    msg = msg + '\n\n' + msg2

                    notification = Notification()
                    notification.member_id = member.pk
                    notification.sender_id = admin.pk
                    notification.message = msg
                    notification.notified_time = str(int(round(time.time() * 1000)))
                    notification.save()

                    rcv = Received()
                    rcv.member_id = member.pk
                    rcv.sender_id = admin.pk
                    rcv.noti_id = notification.pk
                    rcv.save()

                    snt = Sent()
                    snt.member_id = member.pk
                    snt.sender_id = admin.pk
                    snt.noti_id = notification.pk
                    snt.save()

                    ##########################################################################################################################################################################

                    db = firebase.database()
                    data = {
                        "msg": msg,
                        "date":str(int(round(time.time() * 1000))),
                        "sender_id": str(admin.pk),
                        "sender_name": admin.name,
                        "sender_email": admin.email,
                        "sender_photo": admin.photo_url,
                        "role": "admin",
                        "type": "post",
                        "id": str(post.pk),
                        "mes_id": str(notification.pk)
                    }

                    db.child("notify").child(str(member.pk)).push(data)
                    db.child("notify2").child(str(member.pk)).push(data)

                    sendFCMPushNotification(member.pk, admin.pk, msg)

                    #################################################################################################################################################################################

                    if member.playerID != '':
                        playerIDList = []
                        playerIDList.append(member.playerID)
                        url = '/mothers/notifications?noti_id=' + str(notification.pk)
                        send_push(playerIDList, msg, url)

        return redirect('/manager/posts/')



def urlsFromText(str):
    web_urls = re.findall(r'(https?://\S+)', str)
    return web_urls


def createPostUrlPreview(post):
    if post.content != '':
        web_urls = urlsFromText(post.content)
        for wurl in web_urls:
            try:
                preview = link_preview(wurl)
                wtitle = preview.title
                wdescription = preview.description
                wimageurl = preview.image
                wforcetitle = preview.force_title
                wabsoluteimageurl = preview.absolute_image

                icons = favicon.get(wurl)
                icon = None
                if icons is not None and len(icons) > 0: icon = icons[0]

                upreview = PostUrlPreview()
                upreview.post_id = post.pk
                if wtitle is not None: upreview.title = wtitle
                elif wforcetitle is not None: upreview.title = wforcetitle
                if wdescription is not None: upreview.description = wdescription
                if wimageurl is not None and 'http' in wimageurl: upreview.image_url = wimageurl
                elif wabsoluteimageurl is not None: upreview.image_url = wabsoluteimageurl
                if icon is not None: upreview.icon_url = icon.url
                upreview.site_url = wurl
                upreview.save()
            except:
                print('Error')
                try:
                    driver = webdriver.Chrome()
                    driver.get(wurl)
                    wtitle = driver.title

                    icons = favicon.get(wurl)
                    icon = None
                    if icons is not None and len(icons) > 0: icon = icons[0]

                    upreview = PostUrlPreview()
                    upreview.post_id = post.pk
                    if wtitle is not None: upreview.title = wtitle
                    if icon is not None: upreview.icon_url = icon.url
                    upreview.site_url = wurl
                    upreview.save()
                except:
                    pass
            else:
                pass



def add_post_comment(request):

    import datetime

    post_id = request.GET['post_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(admin_id=admin.pk)

    posts = Post.objects.filter(id=post_id)
    if posts.count() == 0: return redirect('/manager/posts/')

    post = posts[0]
    post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")

    pl = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk).first()
    if pl is not None: post.liked = pl.status
    else: post.liked = ''

    comments = Comment.objects.filter(post_id=post.pk, comment_id='0')
    post.comments = str(comments.count())
    likes = PostLike.objects.filter(post_id=post.pk)
    post.reactions = str(likes.count())
    post.content = emoji.emojize(post.content)

    prevs = PostUrlPreview.objects.filter(post_id=post.pk)

    # return HttpResponse(post.member_id + '///' + str(admin.pk))

    if int(post.member_id) != admin.pk:

        pl = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk).first()
        if pl is not None: post.liked = pl.status
        else: post.liked = ''

        ppictures = PostPicture.objects.filter(post_id=post.pk)

        comments = Comment.objects.filter(post_id=post_id, comment_id='0')
        commentList = []
        for comment in comments:
            cl = CommentLike.objects.filter(comment_id=comment.pk, member_id=admin.pk).first()
            if cl is not None: comment.liked = cl.status
            else: comment.liked = ''
            cmts = Comment.objects.filter(comment_id=comment.pk)
            comment.comments = str(cmts.count())
            likes = CommentLike.objects.filter(comment_id=comment.pk)
            comment.reactions = str(likes.count())
            comment.commented_time = datetime.datetime.fromtimestamp(float(int(comment.commented_time)/1000)).strftime("%b %d, %Y %H:%M")
            members = Member.objects.filter(id=comment.member_id)
            if members.count() > 0:
                member = members[0]
                comment.comment_text = emoji.emojize(comment.comment_text)
                data = {
                    'comment':comment,
                    'member':member
                }
                commentList.append(data)
        members = Member.objects.filter(id=post.member_id)
        if members.count() == 0: return redirect('/manager/posts/')
        member = members[0]
        data = {
            'post': post,
            'member': member,
            'pictures':ppictures,
            'prevs': prevs,

        }
        return render(request, 'motherwise/comment.html', {'post':data, 'me':admin, 'comments':commentList})

    else:
        ppictures = PostPicture.objects.filter(post_id=post.pk)
        comments = Comment.objects.filter(post_id=post_id, comment_id='0')
        commentList = []
        for comment in comments:
            cl = CommentLike.objects.filter(comment_id=comment.pk, member_id=admin.pk).first()
            if cl is not None: comment.liked = cl.status
            else: comment.liked = ''
            cmts = Comment.objects.filter(comment_id=comment.pk)
            comment.comments = str(cmts.count())
            likes = CommentLike.objects.filter(comment_id=comment.pk)
            comment.reactions = str(likes.count())
            comment.commented_time = datetime.datetime.fromtimestamp(float(int(comment.commented_time)/1000)).strftime("%b %d, %Y %H:%M")
            members = Member.objects.filter(id=comment.member_id)
            if members.count() > 0:
                member = members[0]
                comment.comment_text = emoji.emojize(comment.comment_text)
                data = {
                    'comment':comment,
                    'member':member
                }
                commentList.append(data)
        data = {
            'post': post,
            'pictures':ppictures,
            'comments':commentList,
            'prevs': prevs,
        }

        categories = []
        pc = PostCategory.objects.filter(admin_id=admin.pk).first()
        if pc is not None:
            if pc.categories != '': categories = pc.categories.split(',')

        return render(request, 'motherwise/edit_post.html', {'post':data, 'me':admin, 'categories':categories})




def delete_post_picture(request):
    picture_id = request.GET['picture_id']
    post_id = request.GET['post_id']
    posts = Post.objects.filter(id=post_id)
    if posts.count() == 0:
        return redirect('/manager/posts/')
    post = posts[0]
    pics = PostPicture.objects.filter(id=picture_id, post_id=post_id)
    fs = FileSystemStorage()
    if pics.count() > 0:
        pic = pics[0]
        if pic.filename != '':
            fs.delete(pic.filename)
        elif pic.picture_url != '':
            fs.delete(pic.picture_url.replace(settings.URL + '/media/', ''))
            if pic.picture_url == post.picture_url:
                post.picture_url = ''
                pics = PostPicture.objects.filter(post_id=post_id)
                if pics.count() > 0:
                    post.picture_url = pics[0].picture_url
                post.save()
        pic.delete()
    return redirect('/manager/add_post_comment?post_id=' + post_id)




def like_post(request):
    post_id = request.GET['post_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    posts = Post.objects.filter(id=post_id)
    if posts.count() == 0: return redirect('/manager/posts/')

    post = posts[0]
    pls = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk)
    if pls.count() > 0:
        pls[0].delete()
        post.likes = str(int(post.likes) - 1)
        post.save()
    else:
        pl = PostLike()
        pl.post_id = post.pk
        pl.member_id = admin.pk
        pl.liked_time = str(int(round(time.time() * 1000)))
        pl.status = 'like'
        pl.save()

        post.likes = str(int(post.likes) + 1)
        post.save()

    # return redirect('/manager/add_post_comment?post_id=' + str(post.pk))
    return HttpResponse(post.likes)




@api_view(['GET', 'POST'])
def submit_comment(request):
    if request.method == 'POST':

        post_id = request.POST.get('post_id', '')
        comment_id = request.POST.get('comment_id', '0')
        content = request.POST.get('content', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return HttpResponse('error')
        except KeyError:
            return HttpResponse('error')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        fs = FileSystemStorage()

        comment = Comment()
        comment.post_id = post_id
        comment.comment_id = comment_id
        comment.member_id = admin.pk
        comment.comment_text = emoji.demojize(content)
        comment.comments = '0'
        comment.likes = '0'
        comment.commented_time = str(int(round(time.time() * 1000)))

        try:
            image = request.FILES['image']
            filename = fs.save(image.name, image)
            uploaded_url = fs.url(filename)
            comment.image_url = settings.URL + uploaded_url
            comment.filename = filename
        except MultiValueDictKeyError:
            print('no video updated')

        comment.save()

        return HttpResponse('success')



def delete_post(request):
    post_id = request.GET['post_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    fs = FileSystemStorage()

    posts = Post.objects.filter(id=post_id)
    if posts.count() == 0: return redirect('/manager/posts/')

    post = posts[0]
    pls = PostLike.objects.filter(post_id=post.pk)
    for pl in pls:
        pl.delete()
    pps = PostPicture.objects.filter(post_id=post.pk)
    for pp in pps:
        if pp.filename != '':
            fs.delete(pp.filename)
        elif pp.picture_url != '':
            fs.delete(pp.picture_url.replace(settings.URL + '/media/', ''))
        pp.delete()
    pcs = Comment.objects.filter(post_id=post.pk)
    for pc in pcs:
        if pc.filename != '':
            fs.delete(pc.filename)
        elif pc.image_url != '':
            fs.delete(pc.image_url.replace(settings.URL + '/media/', ''))
        pc.delete()

    post.delete()

    return redirect('/manager/posts/')



def delete_comment(request):
    comment_id = request.GET['comment_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return HttpResponse('error')
    except KeyError:
        print('no session')
        return HttpResponse('error')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    fs = FileSystemStorage()

    pcs = Comment.objects.filter(id=comment_id)
    if pcs.count() > 0:
        pc = pcs[0]
        post_id = pc.post_id
        if pc.filename != '':
            fs.delete(pc.filename)
        elif pc.image_url != '':
            fs.delete(pc.image_url.replace(settings.URL + '/media/', ''))
        pc.delete()

        return HttpResponse('success')
    else:
        return HttpResponse('error')




@api_view(['GET', 'POST'])
def edit_post(request):
    if request.method == 'POST':

        post_id = request.POST.get('post_id', '1')
        title = request.POST.get('title', '')
        category = request.POST.get('category', '')
        content = request.POST.get('content', '')
        scheduled_time = request.POST.get('scheduled_time', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        posts = Post.objects.filter(id=post_id)
        if posts.count() == 0:
            return redirect('/manager/posts/')
        post = posts[0]
        post.title = title
        post.category = category
        post.content = emoji.demojize(content)
        if post.sch_status != '': post.scheduled_time = scheduled_time
        post.save()

        updatePostUrlPreview(post)

        fs = FileSystemStorage()
        i = 0
        for f in request.FILES.getlist('pictures'):
            i = i + 1
            filename = fs.save(f.name, f)
            uploaded_url = fs.url(filename)
            if i == 1:
                post.picture_url = settings.URL + uploaded_url
                post.save()
            postPicture = PostPicture()
            postPicture.post_id = post.pk
            postPicture.picture_url = settings.URL + uploaded_url
            postPicture.filename = filename
            postPicture.save()

        return redirect('/manager/add_post_comment?post_id=' + post_id)



def updatePostUrlPreview(post):
    if post.content != '':
        web_urls = urlsFromText(post.content)
        for wurl in web_urls:
            try:
                preview = link_preview(wurl)
                wtitle = preview.title
                wdescription = preview.description
                wimageurl = preview.image
                wforcetitle = preview.force_title
                wabsoluteimageurl = preview.absolute_image

                icons = favicon.get(wurl)
                icon = None
                if icons is not None and len(icons) > 0: icon = icons[0]

                upreviews = PostUrlPreview.objects.filter(post_id=post.pk, site_url=wurl)
                if upreviews.count() == 0:
                    upreview = PostUrlPreview()
                    upreview.post_id = post.pk
                    if wtitle is not None: upreview.title = wtitle
                    elif wforcetitle is not None: upreview.title = wforcetitle
                    if wdescription is not None: upreview.description = wdescription
                    if wimageurl is not None and 'http' in wimageurl: upreview.image_url = wimageurl
                    elif wabsoluteimageurl is not None: upreview.image_url = wabsoluteimageurl
                    if icon is not None: upreview.icon_url = icon.url
                    upreview.site_url = wurl
                    upreview.save()
            except:
                print('Error')
                try:
                    driver = webdriver.Chrome()
                    driver.get(wurl)
                    wtitle = driver.title

                    icons = favicon.get(wurl)
                    icon = None
                    if icons is not None and len(icons) > 0: icon = icons[0]

                    upreviews = PostUrlPreview.objects.filter(post_id=post.pk, site_url=wurl)
                    if upreviews.count() == 0:
                        upreview = PostUrlPreview()
                        upreview.post_id = post.pk
                        if wtitle is not None: upreview.title = wtitle
                        if icon is not None: upreview.icon_url = icon.url
                        upreview.site_url = wurl
                        upreview.save()
                except:
                    pass
            else:
                pass


        if len(web_urls) > 0:
            upreviews = PostUrlPreview.objects.filter(post_id=post.pk)
            for upreview in upreviews:
                if not upreview.site_url in web_urls:
                    upreview.delete()
        else:
            upreviews = PostUrlPreview.objects.filter(post_id=post.pk)
            upreviews.delete()

    else:
        upreviews = PostUrlPreview.objects.filter(post_id=post.pk)
        upreviews.delete()





@api_view(['GET', 'POST'])
def search_post(request):

    import datetime

    if request.method == 'POST':
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        uitype = ''
        if request.user_agent.is_mobile:
            uitype = 'mobile'

        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        userList = []
        for user in users:
            if user.registered_time != '':
                user.username = '@' + user.email[0:user.email.find('@')]
                userList.append(user)

        search_id = request.POST.get('q', None)

        posts = Post.objects.filter(sch_status='').order_by('-id')
        postList = get_filtered_posts_data(admin, posts, search_id)

        list1 = []
        list2 = []
        list3 = []
        list4 = []

        i = 0
        itop = 1
        for post in postList:
            post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
            i = i + 1
            pl = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk).first()
            if pl is not None: post.liked = pl.status
            else: post.liked = ''

            comments = Comment.objects.filter(post_id=post.pk, comment_id='0')
            post.comments = str(comments.count())
            likes = PostLike.objects.filter(post_id=post.pk)
            post.reactions = str(likes.count())
            post.content = emoji.emojize(post.content)

            prevs = PostUrlPreview.objects.filter(post_id=post.pk)

            comments1 = comments[:5]
            commentlist = []
            for comment in comments1:
                cm = Member.objects.filter(id=comment.member_id).first()
                if cm is not None:
                    comment.comment_text = emoji.emojize(comment.comment_text)
                    commentlist.append( { 'comment':comment, 'member':cm } )

            member = Member.objects.get(id=post.member_id)

            data = {
                'member':member,
                'post': post,
                'prevs': prevs,
                'comments': commentlist,
                'pc-cnt': str(comments.count()),
            }

            if uitype == 'mobile':
                if 'top' in post.status:
                    list1.insert(0,data)
                else:
                    if i % 4 == 1: list1.append(data)
                    elif i % 4 == 2: list2.append(data)
                    elif i % 4 == 3: list3.append(data)
                    elif i % 4 == 0: list4.append(data)
            else:
                if 'top' in post.status:
                    if itop == 1:
                        list1.insert(0,data)
                        itop = 2
                    elif itop == 2:
                        list2.insert(0,data)
                        itop = 3
                    elif itop == 3:
                        list3.insert(0,data)
                        itop = 4
                    elif itop == 4:
                        list4.insert(0,data)
                        itop = 1
                else:
                    if i % 4 == 1: list1.append(data)
                    elif i % 4 == 2: list2.append(data)
                    elif i % 4 == 3: list3.append(data)
                    elif i % 4 == 0: list4.append(data)

        categories = []
        pc = PostCategory.objects.filter(admin_id=admin.pk).first()
        if pc is not None:
            if pc.categories != '': categories = pc.categories.split(',')

        return render(request, 'motherwise/post.html', {'me':admin, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':'Searched', 'users':userList, 'categories':categories})


# import datetime
from datetime import datetime

def get_filtered_posts_data(me, posts, keyword):
    postList = []
    for post in posts:
        members = Member.objects.filter(id=post.member_id)
        if members.count() > 0:
            member = members[0]
            if int(member.admin_id) == me.pk or int(post.member_id) == me.pk:
                if keyword.lower() in post.title.lower():
                    postList.append(post)
                elif keyword.lower() in post.category.lower():
                    postList.append(post)
                elif keyword.lower() in post.content.lower():
                    postList.append(post)
                elif keyword.lower() in post.comments.lower():
                    postList.append(post)
                elif keyword.lower() in post.reactions.lower():
                    postList.append(post)
                elif keyword.lower() in member.name.lower():
                    postList.append(post)
                elif keyword.lower() in member.email.lower():
                    postList.append(post)
                elif keyword.lower() in member.phone_number.lower():
                    postList.append(post)
                elif keyword.lower() in member.cohort.lower():
                    postList.append(post)
                elif keyword.lower() in member.address.lower():
                    postList.append(post)
                else:
                    if keyword.isdigit():
                        keyDateObj = datetime.fromtimestamp(int(keyword)/1000)
                        postDateObj = datetime.fromtimestamp(int(post.posted_time)/1000)
                        if keyDateObj.year == postDateObj.year and keyDateObj.month == postDateObj.month and keyDateObj.day == postDateObj.day:
                            postList.append(post)

    return postList



def filter(request):

    import datetime

    option = request.GET['option']
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    uitype = ''
    if request.user_agent.is_mobile:
        uitype = 'mobile'

    users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
    userList = []
    for user in users:
        if user.registered_time != '':
            user.username = '@' + user.email[0:user.email.find('@')]
            userList.append(user)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    search = 'Searched'

    allPosts = Post.objects.filter(sch_status='').order_by('-id')
    i = 0
    itop = 1
    for post in allPosts:
        i = i + 1
        pl = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk).first()
        if pl is not None: post.liked = pl.status
        else: post.liked = ''

        comments = Comment.objects.filter(post_id=post.pk, comment_id='0')
        post.comments = str(comments.count())
        likes = PostLike.objects.filter(post_id=post.pk)
        post.reactions = str(likes.count())
        post.content = emoji.emojize(post.content)

        prevs = PostUrlPreview.objects.filter(post_id=post.pk)

        comments1 = comments[:5]
        commentlist = []
        for comment in comments1:
            cm = Member.objects.filter(id=comment.member_id).first()
            if cm is not None:
                comment.comment_text = emoji.emojize(comment.comment_text)
                commentlist.append( { 'comment':comment, 'member':cm } )

        members = Member.objects.filter(id=post.member_id)
        if members.count() > 0:
            memb = members[0]
            if int(memb.admin_id) == admin.pk or memb.pk == admin.pk:
                data = None
                if option == 'last3':
                    if int(round(time.time() * 1000)) - int(post.posted_time) < 3 * 86400 * 1000:
                        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'member':memb,
                            'post': post,
                            'prevs': prevs,
                            'comments': commentlist,
                            'pc-cnt': str(comments.count()),
                        }
                        search = 'Last 3 Days'
                elif option == 'last7':
                    if int(round(time.time() * 1000)) - int(post.posted_time) < 7 * 86400 * 1000:
                        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'member':memb,
                            'post': post,
                            'prevs': prevs,
                            'comments': commentlist,
                            'pc-cnt': str(comments.count()),
                        }
                        search = 'Last 7 Days'
                elif option == 'last30':
                    if int(round(time.time() * 1000)) - int(post.posted_time) < 30 * 86400 * 1000:
                        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'member':memb,
                            'post': post,
                            'prevs': prevs,
                            'comments': commentlist,
                            'pc-cnt': str(comments.count()),
                        }
                        search = 'Last 30 Days'

                if data is not None:
                    if uitype == 'mobile':
                        if 'top' in post.status:
                            list1.insert(0,data)
                        else:
                            if i % 4 == 1: list1.append(data)
                            elif i % 4 == 2: list2.append(data)
                            elif i % 4 == 3: list3.append(data)
                            elif i % 4 == 0: list4.append(data)
                    else:
                        if 'top' in post.status:
                            if itop == 1:
                                list1.insert(0,data)
                                itop = 2
                            elif itop == 2:
                                list2.insert(0,data)
                                itop = 3
                            elif itop == 3:
                                list3.insert(0,data)
                                itop = 4
                            elif itop == 4:
                                list4.insert(0,data)
                                itop = 1
                        else:
                            if i % 4 == 1: list1.append(data)
                            elif i % 4 == 2: list2.append(data)
                            elif i % 4 == 3: list3.append(data)
                            elif i % 4 == 0: list4.append(data)

    categories = []
    pc = PostCategory.objects.filter(admin_id=admin.pk).first()
    if pc is not None:
        if pc.categories != '': categories = pc.categories.split(',')

    return render(request, 'motherwise/post.html', {'me':admin, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':search, 'users':userList, 'categories':categories})



def send_push(playerIDs, message, url):

    client = PybossaOneSignal(api_key=settings.OS_API_KEY, app_id=settings.OS_APP_ID)
    contents = {"en": message}
    headings = {"en": "Motherwise Network"}
    launch_url = settings.URL + url
    chrome_web_image = settings.URL + '/static/images/notimage.jpg'
    chrome_web_icon = settings.URL + '/static/images/noticon.png'
    included_segments = []
    include_player_ids = playerIDs
    web_buttons=[{"id": "read-more-button",
                               "text": "Read more",
                               "icon": "http://i.imgur.com/MIxJp1L.png",
                               "url": launch_url}]
    try:
        client.push_msg(contents=contents, headings=headings, include_player_ids=include_player_ids, launch_url=launch_url, chrome_web_image=chrome_web_image, chrome_web_icon=chrome_web_icon, included_segments=included_segments, web_buttons=web_buttons)
    except:
        print('Error')



def notifications(request):

    import datetime

    noti_id = '0'

    try:
        noti_id = request.GET['noti_id']
    except MultiValueDictKeyError:
        print('No key')

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    notis = Received.objects.filter(member_id=admin.pk).order_by('-id')

    i = 0
    for noti in notis:
        i = i + 1
        members = Member.objects.filter(id=noti.sender_id)
        if members.count() > 0:
            sender = members[0]
            nfs = Notification.objects.filter(id=noti.noti_id)
            if nfs.count() > 0:
                notification = nfs[0]
                notification.notified_time = datetime.datetime.fromtimestamp(float(int(notification.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                data = {
                    'sender':sender,
                    'noti': notification
                }

                if i % 4 == 1: list1.append(data)
                elif i % 4 == 2: list2.append(data)
                elif i % 4 == 3: list3.append(data)
                elif i % 4 == 0: list4.append(data)

    return render(request, 'motherwise/notifications.html', {'notid':noti_id, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'opt':'received'})




def sentnotis(request):

    import datetime

    noti_id = '0'

    try:
        noti_id = request.GET['noti_id']
    except MultiValueDictKeyError:
        print('No key')

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    notis = Sent.objects.filter(sender_id=admin.pk).order_by('-id')

    # return HttpResponse(str(admin.pk) + '///' + str(notis.count()))

    i = 0
    for noti in notis:

        i = i + 1
        members = Member.objects.filter(id=noti.member_id)
        if members.count() > 0:
            receiver = members[0]
            nfs = Notification.objects.filter(id=noti.noti_id)
            if nfs.count() > 0:
                notification = nfs[0]
                notification.notified_time = datetime.datetime.fromtimestamp(float(int(notification.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                data = {
                    'receiver':receiver,
                    'noti': notification
                }

                if i % 4 == 1: list1.append(data)
                elif i % 4 == 2: list2.append(data)
                elif i % 4 == 3: list3.append(data)
                elif i % 4 == 0: list4.append(data)

    return render(request, 'motherwise/sent_notis.html', {'notid':noti_id, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'opt':'sent'})



def delete_noti(request):
    noti_id = request.GET['noti_id']
    opt = request.GET['opt']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    if opt == 'received':
        notis = Received.objects.filter(noti_id=noti_id)
        if notis.count() > 0:
            noti = notis[0]
            noti.delete()
        return redirect('/manager/notifications/')
    elif opt == 'sent':
        notis = Sent.objects.filter(noti_id=noti_id)
        if notis.count() > 0:
            noti = notis[0]
            noti.delete()
        return redirect('/manager/sentnotis/')



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def processnewmessage(request):
    if request.method == 'POST':
        noti_id = request.POST.get('noti_id', '1')
        notis = Notification.objects.filter(id=noti_id)
        if notis.count() > 0:
            noti = notis[0]
            noti.status = 'read'
            noti.save()
        return HttpResponse('The message read')



@api_view(['GET', 'POST'])
def notisearch(request):

    import datetime

    if request.method == 'POST':
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        search_id = request.POST.get('q', None)
        opt = request.POST.get('opt', '')

        notis = []
        if opt == 'received':
            notis = Received.objects.filter(member_id=admin.pk).order_by('-id')
        elif opt == 'sent':
            notis = Sent.objects.filter(sender_id=admin.pk).order_by('-id')
        notiList = []
        for noti in notis:
            nfs = Notification.objects.filter(id=noti.noti_id)
            if nfs.count() > 0:
                notification = nfs[0]
                notiList.append(notification)
        notiList = get_filtered_notis_data(notiList, search_id, opt)

        list1 = []
        list2 = []
        list3 = []
        list4 = []

        i = 0
        for noti in notiList:
            noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
            i = i + 1
            members = []
            if opt == 'received':
                members = Member.objects.filter(id=noti.sender_id)
            elif opt == 'sent':
                members = Member.objects.filter(id=noti.member_id)
            if members.count() > 0:
                member = members[0]
                data = {}
                if opt == 'received':
                    data = {
                        'sender':member,
                        'noti': noti
                    }
                elif opt == 'sent':
                    data = {
                        'receiver':member,
                        'noti': noti
                    }

                if i % 4 == 1: list1.append(data)
                elif i % 4 == 2: list2.append(data)
                elif i % 4 == 3: list3.append(data)
                elif i % 4 == 0: list4.append(data)

        if opt == 'sent':
            return render(request, 'motherwise/sent_notis.html', {'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':'Searched', 'opt':'sent'})

        return render(request, 'motherwise/notifications.html', {'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':'Searched', 'opt':'received'})


# import datetime
# from datetime import datetime

def get_filtered_notis_data(notis, keyword, option):
    notiList = []
    for noti in notis:
        members = []
        if option == 'received':
            members = Member.objects.filter(id=noti.sender_id)
        elif option == 'sent':
            members = Member.objects.filter(id=noti.member_id)
        if members.count() > 0:
            member = members[0]
            if keyword.lower() in noti.message.lower():
                notiList.append(noti)
            elif keyword.lower() in member.name.lower():
                notiList.append(noti)
            elif keyword.lower() in member.email.lower():
                notiList.append(noti)
            elif keyword.lower() in member.phone_number.lower():
                notiList.append(noti)
            elif keyword.lower() in member.cohort.lower():
                notiList.append(noti)
            elif keyword.lower() in member.address.lower():
                notiList.append(noti)
            else:
                if keyword.isdigit():
                    keyDateObj = datetime.fromtimestamp(int(keyword)/1000)
                    notiDateObj = datetime.fromtimestamp(int(noti.notified_time)/1000)
                    if keyDateObj.year == notiDateObj.year and keyDateObj.month == notiDateObj.month and keyDateObj.day == notiDateObj.day:
                        notiList.append(noti)

    return notiList



def fffff(request):

    import datetime

    option = request.GET['noption']
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    search = 'Searched'
    opt = request.GET['opt']

    notis = []
    if opt == 'received':
        notis = Received.objects.filter(member_id=admin.pk).order_by('-id')
    elif opt == 'sent':
        notis = Sent.objects.filter(sender_id=admin.pk).order_by('-id')

    notiList = []
    for noti in notis:
        nfs = Notification.objects.filter(id=noti.noti_id)
        if nfs.count() > 0:
            notification = nfs[0]
            notiList.append(notification)

    i = 0
    for noti in notiList:
        i = i + 1
        if opt == 'received':
            members = Member.objects.filter(id=noti.sender_id)
            if members.count() > 0:
                sender = members[0]
                if option == 'last3':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 3 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'sender':sender,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 3 Days'
                elif option == 'last7':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 7 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'sender':sender,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 7 Days'
                elif option == 'last30':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 30 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'sender':sender,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 30 Days'
        elif opt == 'sent':
            members = Member.objects.filter(id=noti.member_id)
            if members.count() > 0:
                receiver = members[0]
                if option == 'last3':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 3 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'receiver':receiver,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 3 Days'
                elif option == 'last7':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 7 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'receiver':receiver,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 7 Days'
                elif option == 'last30':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 30 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'receiver':receiver,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 30 Days'

    if opt == 'sent':
        return render(request, 'motherwise/sent_notis.html', {'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':'Searched', 'opt':'sent'})

    return render(request, 'motherwise/notifications.html', {'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':search, 'opt':'received'})




def videotest(request):
    return render(request, 'motherwise/video_agora.html')


def open_conference(request):

    import datetime

    group_id = request.GET['group_id']
    # cohort = request.GET['cohort']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    memberList = []
    if group_id != '' and int(group_id) > 0:
        group = None
        groups = Group.objects.filter(member_id=admin.pk, id=group_id).order_by('-id')
        if groups.count() > 0:
            group = groups[0]
            gMembers = GroupMember.objects.filter(group_id=group.pk)
            for gMember in gMembers:
                members = Member.objects.filter(id=gMember.member_id)
                if members.count() > 0:
                    memberList.append(members[0])

            request.session['group_id'] = group_id
            request.session['cohort'] = ''

            memberIdList = []
            for memb in memberList:
                memberIdList.append(memb.pk)
            request.session['selected_member_list'] = memberIdList

            confs = Conference.objects.filter(member_id=admin.pk, group_id=group_id).order_by('-id')
            for conf in confs:
                conf.created_time = datetime.datetime.fromtimestamp(float(int(conf.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                if conf.event_time != '': conf.event_time = datetime.datetime.fromtimestamp(float(int(conf.event_time)/1000)).strftime("%b %d, %Y %H:%M")
            if confs.count() == 0:
                return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'group':group})
            try:
                option = request.GET['option']
                if option == 'new_conference':
                    return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'group':group})
            except KeyError:
                print('no key')
            # return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'group':group})

            last_conf = confs[0]

            try:
                conf_id = request.GET['conf_id']
                cfs = Conference.objects.filter(id=conf_id)
                if cfs.count() > 0:
                    last_conf = cfs[0]
                    last_conf.created_time = datetime.datetime.fromtimestamp(float(int(last_conf.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    if last_conf.event_time != '': last_conf.event_time = datetime.datetime.fromtimestamp(float(int(last_conf.event_time)/1000)).strftime("%b %d, %Y %H:%M")
            except KeyError:
                print('no key')

            if last_conf.type == 'live':
                return render(request, 'motherwise/conference_live.html', {'me':admin, 'members':memberList, 'group':group, 'confs':confs, 'last_conf':last_conf})
            elif last_conf.type == 'file':
                return render(request, 'motherwise/conference_video.html', {'me':admin, 'members':memberList, 'group':group, 'confs':confs, 'last_conf':last_conf})
            elif last_conf.type == 'youtube':
                return render(request, 'motherwise/conference_youtube.html', {'me':admin, 'members':memberList, 'group':group, 'confs':confs, 'last_conf':last_conf})
        else:
            return redirect('/manager/home')

    else:
        memberList = Member.objects.filter(admin_id=admin.pk, status='')

        request.session['group_id'] = ''
        request.session['cohort'] = ''

        memberIdList = []
        for memb in memberList:
            memberIdList.append(memb.pk)
        request.session['selected_member_list'] = memberIdList

        confs = Conference.objects.filter(member_id=admin.pk, group_id=0).order_by('-id')
        for conf in confs:
            conf.created_time = datetime.datetime.fromtimestamp(float(int(conf.created_time)/1000)).strftime("%b %d, %Y %H:%M")
            if conf.event_time != '': conf.event_time = datetime.datetime.fromtimestamp(float(int(conf.event_time)/1000)).strftime("%b %d, %Y %H:%M")
        if confs.count() == 0:
            return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'group':None})
        try:
            option = request.GET['option']
            if option == 'new_conference':
                return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'group':None})
        except KeyError:
            print('no key')
        # return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'group':None})

        last_conf = confs[0]

        try:
            conf_id = request.GET['conf_id']
            cfs = Conference.objects.filter(id=conf_id)
            if cfs.count() > 0:
                last_conf = cfs[0]
                last_conf.created_time = datetime.datetime.fromtimestamp(float(int(last_conf.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                if last_conf.event_time != '': last_conf.event_time = datetime.datetime.fromtimestamp(float(int(last_conf.event_time)/1000)).strftime("%b %d, %Y %H:%M")
        except KeyError:
            print('no key')

        if last_conf.type == 'live':
            return render(request, 'motherwise/conference_live.html', {'me':admin, 'members':memberList, 'group':None, 'confs':confs, 'last_conf':last_conf})
        elif last_conf.type == 'file':
            return render(request, 'motherwise/conference_video.html', {'me':admin, 'members':memberList, 'group':None, 'confs':confs, 'last_conf':last_conf})
        elif last_conf.type == 'youtube':
            return render(request, 'motherwise/conference_youtube.html', {'me':admin, 'members':memberList, 'group':None, 'confs':confs, 'last_conf':last_conf})



def genRandomConferenceCode():
    import strgen
    randomString = strgen.StringGenerator("[\w\d]{6}").render()
    return randomString



@api_view(['GET', 'POST'])
def create_conference(request):

    import datetime

    if request.method == 'POST':

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        group_id = request.POST.get('group_id', '1')
        group_option = request.POST.get('group_option', '')

        name = request.POST.get('name', '')
        type = request.POST.get('type', '')
        youtubeurl = request.POST.get('youtubeurl', '')

        conf = Conference()
        conf.member_id = admin.pk
        if group_option == 'community':
            conf.group_id = group_id
        else:
            group_id = '0'
            conf.group_id = group_id
        conf.name = name
        if type == 'live':
            conf.code = genRandomConferenceCode()
        conf.type = type
        # conf.event_time = event_time
        conf.created_time = str(int(round(time.time() * 1000)))
        conf.participants = '0'
        conf.likes = '0'

        dt = ''

        if type == 'youtube':
            conf.video_url = youtubeurl
            # if event_time != '':
            #     dt = datetime.datetime.fromtimestamp(float(int(event_time)/1000)).strftime("%b %d, %Y %H:%M")
        elif type == 'file':
            try:
                videofile = request.FILES['video']
                fs = FileSystemStorage()
                filename = fs.save(videofile.name, videofile)
                video_url = fs.url(filename)
                conf.video_url = settings.URL + video_url
                conf.filename = filename
            except MultiValueDictKeyError:
                print("File Not Exist")

            # if event_time != '':
            #     dt = datetime.datetime.fromtimestamp(float(int(event_time)/1000)).strftime("%b %d, %Y %H:%M")

        # elif type == 'live':
        #     conf.video_url = ''
            # dt = datetime.datetime.fromtimestamp(float(int(event_time)/1000)).strftime("%b %d, %Y %H:%M")

        conf.save()

        option = request.POST.get('option')

        if option == 'message':
            if group_id != '' and int(group_id) > 0:
                groups = Group.objects.filter(id=group_id)
                if groups.count() > 0:
                    group = groups[0]
                    gMembers = GroupMember.objects.filter(group_id=group.pk)
                    for gMember in gMembers:
                        members = Member.objects.filter(id=gMember.member_id)
                        if members.count() > 0:
                            member = members[0]

                            notification = Notification()
                            notification.member_id = member.pk
                            notification.sender_id = admin.pk
                            notification.notified_time = str(int(round(time.time() * 1000)))
                            notification.save()

                            rcv = Received()
                            rcv.member_id = member.pk
                            rcv.sender_id = admin.pk
                            rcv.noti_id = notification.pk
                            rcv.save()

                            snt = Sent()
                            snt.member_id = member.pk
                            snt.sender_id = admin.pk
                            snt.noti_id = notification.pk
                            snt.save()

                            title = 'Video Conference Invitation In The Nest'
                            subject = 'You\'ve received an invitation in the Nest. (has recibido una invitación en el Nest.)'
                            msg = 'Dear ' + member.name + ',<br><br>You\'ve received an invitation to a video conference in a community ' + group.name + ' from MotherWise Community: The Nest.<br><br>Community name: ' + group.name + '<br>'
                            msg = msg + 'Conference topic: ' + name + '<br>'
                            msg = msg + 'Entry code: ' + conf.code + '<br>'
                            if type == 'live': type = 'Live'
                            elif type == 'file': type = 'General'
                            elif type == 'youtube': type = 'YouTube'
                            msg = msg + 'Type: ' + type + '<br>'
                            # if dt != '':
                            #     msg = msg + 'Conference start date and time: ' + dt + '<br>'
                            msg = msg + 'So you can see this conference in the community and connect it to attend the event at that time.<br>'
                            msg = msg + 'Please join the community conference.'
                            msg = msg + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>MotherWise Team'

                            title2 = 'invitación de videoconferencia en el Nest'
                            msg2 = member.name + ',<br><br>has recibido una invitación a una videoconferencia en una comunidad ' + group.name + ' de la comunidad MotherWise: el Nest.<br><br>Nombre de la comunidad: ' + group.name + '<br>'
                            msg2 = msg2 + 'tema de la conferencia: ' + name + '<br>'
                            msg2 = msg2 + 'código de entrada: ' + conf.code + '<br>'
                            if type == 'live': type = 'en vivo'
                            elif type == 'file': type = 'general'
                            elif type == 'youtube': type = 'YouTube'
                            msg2 = msg2 + 'Type: ' + type + '<br>'
                            # if dt != '':
                            #     msg2 = msg2 + 'fecha y hora de inicio de la conferencia: ' + dt + '<br>'
                            msg2 = msg2 + 'para que pueda ver esta conferencia en la comunidad y conectarla para asistir al evento en ese momento.<br>'
                            msg2 = msg2 + 'por favor únase a la conferencia comunitaria.'
                            msg2 = msg2 + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>unirse al sitio web</a>' + '<br><br>Equipo MotherWise'

                            from_email = admin.email
                            to_emails = []
                            to_emails.append(member.email)
                            send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                            msg = 'Dear ' + member.name + ',\nYou\'ve received an invitation to a community ' + group.name + ' from MotherWise Community: The Nest.\nCommunity name: ' + group.name + '\n'
                            msg = msg + 'Conference topic: ' + name + '\n'
                            msg = msg + 'Entry code: ' + conf.code + '\n'
                            if type == 'live': type = 'Live'
                            elif type == 'file': type = 'General'
                            elif type == 'youtube': type = 'YouTube'
                            msg = msg + 'Type: ' + type + '\n'
                            # if dt != '':
                            #     msg = msg + 'Conference start date and time: ' + dt + '\n'
                            msg = msg + 'So you can see this conference in the community and connect it to attend the event at that time.\n'
                            msg = msg + 'Please join the community conference.\nBest Regards\n\nMotherWise Community'

                            msg2 = member.name + ',\nhas recibido una invitación a una comunidad ' + group.name + ' de la comunidad MotherWise: el Nest.\nnombre de la comunidad: ' + group.name + '\n'
                            msg2 = msg2 + 'tema de la conferencia: ' + name + '\n'
                            msg2 = msg2 + 'código de entrada: ' + conf.code + '\n'
                            if type == 'live': type = 'en vivo'
                            elif type == 'file': type = 'general'
                            elif type == 'youtube': type = 'YouTube'
                            msg2 = msg2 + 'Type: ' + type + '\n'
                            # if dt != '':
                            #     msg2 = msg2 + 'fecha y hora de inicio de la conferencia: ' + dt + '\n'
                            msg2 = msg2 + 'para que pueda ver esta conferencia en la comunidad y conectarla para asistir al evento en ese momento.\n'
                            msg2 = msg2 + 'por favor únase a la conferencia comunitaria.\natentamente\n\ncomunidad MotherWise'

                            msg = msg + '\n\n' + msg2

                            notification.message = msg
                            notification.save()

                            ##########################################################################################################################################################################

                            db = firebase.database()
                            data = {
                                "msg": msg,
                                "date":str(int(round(time.time() * 1000))),
                                "sender_id": str(admin.pk),
                                "sender_name": admin.name,
                                "sender_email": admin.email,
                                "sender_photo": admin.photo_url,
                                "role": "admin",
                                "type": "conference",
                                "id": str(conf.pk),
                                "mes_id": str(notification.pk)
                            }

                            db.child("notify").child(str(member.pk)).push(data)
                            db.child("notify2").child(str(member.pk)).push(data)

                            sendFCMPushNotification(member.pk, admin.pk, msg)

                            #################################################################################################################################################################################

                            if member.playerID != '':
                                playerIDList = []
                                playerIDList.append(member.playerID)
                                url = '/mothers/notifications?noti_id=' + str(notification.pk)
                                send_push(playerIDList, msg, url)

                    conf.status = 'notified'
                    conf.save()

            else:
                members = Member.objects.filter(admin_id=admin.pk)
                if members.count() > 0:
                    for member in members:

                        notification = Notification()
                        notification.member_id = member.pk
                        notification.sender_id = admin.pk
                        notification.notified_time = str(int(round(time.time() * 1000)))
                        notification.save()

                        rcv = Received()
                        rcv.member_id = member.pk
                        rcv.sender_id = admin.pk
                        rcv.noti_id = notification.pk
                        rcv.save()

                        snt = Sent()
                        snt.member_id = member.pk
                        snt.sender_id = admin.pk
                        snt.noti_id = notification.pk
                        snt.save()

                        title = 'Video Conference Invitation In The Nest'
                        subject = 'You\'ve received an invitation in the Nest. (has recibido una invitación en el Nest.)'
                        msg = 'Dear ' + member.name + ',<br><br>You\'ve received an invitation to a video conference from MotherWise Community: The Nest.<br><br>'
                        msg = msg + 'Conference topic: ' + conf.name + '<br>'
                        msg = msg + 'Entry code: ' + conf.code + '<br>'
                        type = conf.type
                        if type == 'live': type = 'Live'
                        elif type == 'file': type = 'General'
                        elif type == 'youtube': type = 'YouTube'
                        msg = msg + 'Type: ' + type + '<br>'
                        # if dt != '':
                        #     msg = msg + 'Conference start date and time: ' + dt + '<br>'
                        msg = msg + 'So you can see this conference in the group and connect it to attend the event at that time.<br>'
                        msg = msg + 'Please join the group conference.'
                        msg = msg + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>MotherWise Team'

                        title2 = 'invitación de videoconferencia In the Nest'
                        msg2 = member.name + ',<br><br>has recibido una invitación a una videoconferencia de la comunidad MotherWise: el Nest.<br><br>'
                        msg2 = msg2 + 'tema de la conferencia: ' + conf.name + '<br>'
                        msg2 = msg2 + 'código de entrada: ' + conf.code + '<br>'
                        type = conf.type
                        if type == 'live': type = 'en vivo'
                        elif type == 'file': type = 'general'
                        elif type == 'youtube': type = 'YouTube'
                        msg2 = msg2 + 'Type: ' + type + '<br>'
                        # if dt != '':
                        #     msg2 = msg2 + 'fecha y hora de inicio de la conferencia: ' + dt + '<br>'
                        msg2 = msg2 + 'para que pueda ver esta conferencia en el grupo y conectarla para asistir al evento en ese momento.<br>'
                        msg2 = msg2 + 'por favor únase a la conferencia grupal.'
                        msg2 = msg2 + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>unirse al sitio web</a>' + '<br><br>Equipo MotherWise'

                        from_email = admin.email
                        to_emails = []
                        to_emails.append(member.email)
                        send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                        msg = 'Dear ' + member.name + ',\nYou\'ve received an invitation from MotherWise Community: The Nest.\n'
                        msg = msg + 'Conference topic: ' + conf.name + '\n'
                        msg = msg + 'Entry code: ' + conf.code + '\n'
                        type = conf.type
                        if type == 'live': type = 'Live'
                        elif type == 'file': type = 'General'
                        elif type == 'youtube': type = 'YouTube'
                        msg = msg + 'Type: ' + type + '\n'
                        # if dt != '':
                        #     msg = msg + 'Conference start date and time: ' + dt + '\n'
                        msg = msg + 'So you can see this conference in the group and connect it to attend the event at that time.\n'
                        msg = msg + 'Please join the group conference.\nBest Regards\n\nMotherWise Community'

                        msg2 = member.name + ',\nhas recibido una invitación a una videoconferencia de la comunidad MotherWise: el Nest.\n'
                        msg2 = msg2 + 'tema de la conferencia: ' + conf.name + '\n'
                        msg2 = msg2 + 'código de entrada: ' + conf.code + '\n'
                        type = conf.type
                        if type == 'live': type = 'en vivo'
                        elif type == 'file': type = 'general'
                        elif type == 'youtube': type = 'YouTube'
                        msg2 = msg2 + 'Type: ' + type + '\n'
                        # if dt != '':
                        #     msg2 = msg2 + 'fecha y hora de inicio de la conferencia: ' + dt + '\n'
                        msg2 = msg2 + 'para que pueda ver esta conferencia en el grupo y conectarla para asistir al evento en ese momento.\n'
                        msg2 = msg2 + 'por favor únase a la conferencia grupal.\n\nComunidad MotherWise'

                        msg = msg + '\n\n' + msg2

                        notification.message = msg
                        notification.save()

                        ##########################################################################################################################################################################

                        db = firebase.database()
                        data = {
                            "msg": msg,
                            "date":str(int(round(time.time() * 1000))),
                            "sender_id": str(admin.pk),
                            "sender_name": admin.name,
                            "sender_email": admin.email,
                            "sender_photo": admin.photo_url,
                            "role": "admin",
                            "type": "conference",
                            "id": str(conf.pk),
                            "mes_id": str(notification.pk)
                        }

                        db.child("notify").child(str(member.pk)).push(data)
                        db.child("notify2").child(str(member.pk)).push(data)

                        sendFCMPushNotification(member.pk, admin.pk, msg)

                        #################################################################################################################################################################################

                        if member.playerID != '':
                            playerIDList = []
                            playerIDList.append(member.playerID)
                            url = '/mothers/notifications?noti_id=' + str(notification.pk)
                            send_push(playerIDList, msg, url)

                    conf.status = 'notified'
                    conf.save()


        else:
            print('no message')

        try:
            branch = request.POST.get('branch', '')
            if branch == 'conferences':
                return redirect('/manager/to_conferences')
        except KeyError:
            print('no key')

        return redirect('/manager/open_conference?conf_id=' + str(conf.pk) + '&group_id=' + group_id)




@api_view(['GET', 'POST'])
def delete_conference(request):
    if request.method == 'GET':
        conf_id = request.GET['conf_id']

        confs = Conference.objects.filter(id=conf_id)

        fs = FileSystemStorage()

        if confs.count() > 0:
            conf = confs[0]
            group_id = conf.group_id
            cohort = conf.cohort
            if conf.filename != '':
                fs.delete(conf.filename)
            elif conf.video_url != '' and 'http' in conf.video_url:
                fname = conf.video_url.replace(settings.URL + '/media/', '')
                fs.delete(fname)
            conf.delete()

            try:
                branch = request.GET['branch']
                if branch == 'conferences':
                    return redirect('/manager/to_conferences')
            except KeyError:
                print('no key')

            return redirect('/manager/open_conference?group_id=' + group_id)
        else:
            return redirect('/manager/home')




@api_view(['GET', 'POST'])
def conference_notify(request):

    import datetime

    if request.method == 'POST':

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        group_id = request.POST.get('group_id', '1')
        conf_id = request.POST.get('conf_id', '1')

        confs = Conference.objects.filter(id=conf_id)
        if confs.count() == 0:
            return redirect('/manager/home')

        conf = confs[0]

        name = conf.name
        type = conf.type

        dt = ''

        # if event_time != '':
        #     dt = datetime.datetime.fromtimestamp(float(int(event_time)/1000)).strftime("%b %d, %Y %H:%M")

        if admin is not None:
            if group_id != '' and int(group_id) > 0:
                groups = Group.objects.filter(id=group_id)
                if groups.count() > 0:
                    group = groups[0]
                    gMembers = GroupMember.objects.filter(group_id=group.pk)
                    for gMember in gMembers:
                        members = Member.objects.filter(id=gMember.member_id)
                        if members.count() > 0:
                            member = members[0]

                            notification = Notification()
                            notification.member_id = member.pk
                            notification.sender_id = admin.pk
                            notification.notified_time = str(int(round(time.time() * 1000)))
                            notification.save()

                            rcv = Received()
                            rcv.member_id = member.pk
                            rcv.sender_id = admin.pk
                            rcv.noti_id = notification.pk
                            rcv.save()

                            snt = Sent()
                            snt.member_id = member.pk
                            snt.sender_id = admin.pk
                            snt.noti_id = notification.pk
                            snt.save()

                            title = 'Video Conference Invitation In The Nest'
                            subject = 'You\'ve received an invitation in the Nest (has recibido una invitación en el Nest.)'
                            msg = 'Dear ' + member.name + ',<br><br>You\'ve received an invitation to a video conference in a community ' + group.name + ' from MotherWise Community: The Nest.<br><br>Community name: ' + group.name + '<br>'
                            msg = msg + 'Conference topic: ' + conf.name + '<br>'
                            msg = msg + 'Entry code: ' + conf.code + '<br>'
                            type = conf.type
                            if type == 'live': type = 'Live'
                            elif type == 'file': type = 'General'
                            elif type == 'youtube': type = 'YouTube'
                            msg = msg + 'Type: ' + type + '<br>'
                            # if dt != '':
                            #     msg = msg + 'Conference start date and time: ' + dt + '<br>'
                            msg = msg + 'So you can see this conference in the community and connect it to attend the event at that time.<br>'
                            msg = msg + 'Please join the community conference.'
                            msg = msg + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>MotherWise Team'

                            title2 = 'invitación de videoconferencia In the Nest'
                            msg2 = member.name + ',<br><br>has recibido una invitación a una videoconferencia en una comunidad ' + group.name + ' de la comunidad MotherWise: el Nest.<br><br>nombre de la comunidad: ' + group.name + '<br>'
                            msg2 = msg2 + 'tema de la conferencia: ' + conf.name + '<br>'
                            msg2 = msg2 + 'código de entrada: ' + conf.code + '<br>'
                            type = conf.type
                            if type == 'live': type = 'en vivo'
                            elif type == 'file': type = 'general'
                            elif type == 'youtube': type = 'YouTube'
                            msg2 = msg2 + 'Type: ' + type + '<br>'
                            # if dt != '':
                            #     msg2 = msg2 + 'fecha y hora de inicio de la conferencia: ' + dt + '<br>'
                            msg2 = msg2 + 'para que pueda ver esta conferencia en la comunidad y conectarla para asistir al evento en ese momento.<br>'
                            msg2 = msg2 + 'por favor únase a la conferencia comunitaria.'
                            msg2 = msg2 + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>unirse al sitio web</a>' + '<br><br>Equipo MotherWise'

                            from_email = admin.email
                            to_emails = []
                            to_emails.append(member.email)
                            send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                            msg = 'Dear ' + member.name + ',\nYou\'ve received an invitation to a community ' + group.name + ' from MotherWise Community: The Nest.\nCommunity name: ' + group.name + '\n'
                            msg = msg + 'Conference topic: ' + conf.name + '\n'
                            msg = msg + 'Entry code: ' + conf.code + '\n'
                            type = conf.type
                            if type == 'live': type = 'Live'
                            elif type == 'file': type = 'General'
                            elif type == 'youtube': type = 'YouTube'
                            msg = msg + 'Type: ' + type + '\n'
                            # if dt != '':
                            #     msg = msg + 'Conference start date and time: ' + dt + '\n'
                            msg = msg + 'So you can see this conference in the community and connect it to attend the event at that time.\n'
                            msg = msg + 'Please join the community conference.\nBest Regards\n\nMotherWise Community'

                            msg2 = member.name + ',\nhas recibido una invitación a una videoconferencia en una comunidad ' + group.name + ' de la comunidad MotherWise: el Nest.\nnombre de la comunidad: ' + group.name + '\n'
                            msg2 = msg2 + 'tema de la conferencia: ' + conf.name + '\n'
                            msg2 = msg2 + 'código de entrada: ' + conf.code + '\n'
                            type = conf.type
                            if type == 'live': type = 'en vivo'
                            elif type == 'file': type = 'general'
                            elif type == 'youtube': type = 'YouTube'
                            msg2 = msg2 + 'Type: ' + type + '\n'
                            # if dt != '':
                            #     msg2 = msg2 + 'fecha y hora de inicio de la conferencia: ' + dt + '\n'
                            msg2 = msg2 + 'para que pueda ver esta conferencia en la comunidad y conectarla para asistir al evento en ese momento.\n'
                            msg2 = msg2 + 'por favor únase a la conferencia comunitaria.\n\nComunidad MotherWise'

                            msg = msg + '\n\n' + msg2

                            notification.message = msg
                            notification.save()

                            ##########################################################################################################################################################################

                            db = firebase.database()
                            data = {
                                "msg": msg,
                                "date":str(int(round(time.time() * 1000))),
                                "sender_id": str(admin.pk),
                                "sender_name": admin.name,
                                "sender_email": admin.email,
                                "sender_photo": admin.photo_url,
                                "role": "admin",
                                "type": "conference",
                                "id": str(conf.pk),
                                "mes_id": str(notification.pk)
                            }

                            db.child("notify").child(str(member.pk)).push(data)
                            db.child("notify2").child(str(member.pk)).push(data)

                            sendFCMPushNotification(member.pk, admin.pk, msg)

                            #################################################################################################################################################################################

                            if member.playerID != '':
                                playerIDList = []
                                playerIDList.append(member.playerID)
                                url = '/mothers/notifications?noti_id=' + str(notification.pk)
                                send_push(playerIDList, msg, url)

                    conf.status = 'notified'
                    conf.save()

            else:
                members = Member.objects.filter(admin_id=admin.pk)
                if members.count() > 0:
                    for member in members:

                        notification = Notification()
                        notification.member_id = member.pk
                        notification.sender_id = admin.pk
                        notification.notified_time = str(int(round(time.time() * 1000)))
                        notification.save()

                        rcv = Received()
                        rcv.member_id = member.pk
                        rcv.sender_id = admin.pk
                        rcv.noti_id = notification.pk
                        rcv.save()

                        snt = Sent()
                        snt.member_id = member.pk
                        snt.sender_id = admin.pk
                        snt.noti_id = notification.pk
                        snt.save()

                        title = 'Video Conference Invitation In The Nest'
                        subject = 'You\'ve received an invitation in the Nest. (has recibido una invitación en el Nest.)'
                        msg = 'Dear ' + member.name + ',<br><br>You\'ve received an invitation to a video conference from MotherWise Community: The Nest.<br><br>'
                        msg = msg + 'Conference topic: ' + conf.name + '<br>'
                        msg = msg + 'Entry code: ' + conf.code + '<br>'
                        type = conf.type
                        if type == 'live': type = 'Live'
                        elif type == 'file': type = 'General'
                        elif type == 'youtube': type = 'YouTube'
                        msg = msg + 'Type: ' + type + '<br>'
                        # if dt != '':
                        #     msg = msg + 'Conference start date and time: ' + dt + '<br>'
                        msg = msg + 'So you can see this conference in the group and connect it to attend the event at that time.<br>'
                        msg = msg + 'Please join the group conference.'
                        msg = msg + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>MotherWise Team'

                        title2 = 'invitación de videoconferencia In the Nest'
                        msg2 = member.name + ',<br><br>has recibido una invitación a una videoconferencia de la comunidad MotherWise: el Nest.<br><br>'
                        msg2 = msg2 + 'tema de la conferencia: ' + conf.name + '<br>'
                        msg2 = msg2 + 'código de entrada: ' + conf.code + '<br>'
                        type = conf.type
                        if type == 'live': type = 'en vivo'
                        elif type == 'file': type = 'general'
                        elif type == 'youtube': type = 'YouTube'
                        msg2 = msg2 + 'Type: ' + type + '<br>'
                        # if dt != '':
                        #     msg2 = msg2 + 'fecha y hora de inicio de la conferencia: ' + dt + '<br>'
                        msg2 = msg2 + 'para que pueda ver esta conferencia en el grupo y conectarla para asistir al evento en ese momento.<br>'
                        msg2 = msg2 + 'por favor únase a la conferencia grupal.'
                        msg2 = msg2 + '<br><a href=\'' + settings.URL + '/mothers/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>unirse al sitio web</a>' + '<br><br>Equipo MotherWise'

                        from_email = admin.email
                        to_emails = []
                        to_emails.append(member.email)
                        send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                        msg = 'Dear ' + member.name + ',\nYou\'ve received an invitation from MotherWise Community: The Nest.\n'
                        msg = msg + 'Conference topic: ' + conf.name + '\n'
                        msg = msg + 'Entry code: ' + conf.code + '\n'
                        type = conf.type
                        if type == 'live': type = 'Live'
                        elif type == 'file': type = 'General'
                        elif type == 'youtube': type = 'YouTube'
                        msg = msg + 'Type: ' + type + '\n'
                        # if dt != '':
                        #     msg = msg + 'Conference start date and time: ' + dt + '\n'
                        msg = msg + 'So you can see this conference in the group and connect it to attend the event at that time.\n'
                        msg = msg + 'Please join the group conference.\nBest Regards\n\nMotherWise Community'

                        msg2 = member.name + ',\nhas recibido una invitación a una videoconferencia de la comunidad MotherWise: el Nest.\n'
                        msg2 = msg2 + 'tema de la conferencia: ' + conf.name + '\n'
                        msg2 = msg2 + 'código de entrada: ' + conf.code + '\n'
                        type = conf.type
                        if type == 'live': type = 'en vivo'
                        elif type == 'file': type = 'general'
                        elif type == 'youtube': type = 'YouTube'
                        msg2 = msg2 + 'Type: ' + type + '\n'
                        # if dt != '':
                        #     msg2 = msg2 + 'fecha y hora de inicio de la conferencia: ' + dt + '\n'
                        msg2 = msg2 + 'para que pueda ver esta conferencia en el grupo y conectarla para asistir al evento en ese momento.\n'
                        msg2 = msg2 + 'por favor únase a la conferencia grupal.\n\nComunidad MotherWise'

                        msg = msg + '\n\n' + msg2

                        notification.message = msg
                        notification.save()

                        ##########################################################################################################################################################################

                        db = firebase.database()
                        data = {
                            "msg": msg,
                            "date":str(int(round(time.time() * 1000))),
                            "sender_id": str(admin.pk),
                            "sender_name": admin.name,
                            "sender_email": admin.email,
                            "sender_photo": admin.photo_url,
                            "role": "admin",
                            "type": "conference",
                            "id": str(conf.pk),
                            "mes_id": str(notification.pk)
                        }

                        db.child("notify").child(str(member.pk)).push(data)
                        db.child("notify2").child(str(member.pk)).push(data)

                        sendFCMPushNotification(member.pk, admin.pk, msg)

                        #################################################################################################################################################################################

                        if member.playerID != '':
                            playerIDList = []
                            playerIDList.append(member.playerID)
                            url = '/mothers/notifications?noti_id=' + str(notification.pk)
                            send_push(playerIDList, msg, url)

                    conf.status = 'notified'
                    conf.save()

        else:
            return render(request, 'motherwise/admin.html')

        return redirect('/manager/open_conference?conf_id=' + str(conf.pk) + '&group_id=' + group_id + '&cohort=' + cohort)




@api_view(['GET', 'POST'])
def video_selected_members(request):

    import datetime

    if request.method == 'POST':

        ids = request.POST.getlist('users[]')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        memberList = []
        memberList2 = []
        memberIdList = []
        for member_id in ids:
            members = Member.objects.filter(id=int(member_id))
            if members.count() > 0:
                member = members[0]
                memberList.append(member)
                memberIdList.append(member.pk)

        if len(memberList) > 0:
            request.session['selected_member_list'] = memberIdList
            groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
            for group in groups:
                group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
            latestGroupMemberList = []
            latest_group = None
            if groups.count() > 0:
                latest_group = groups[0]
                gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        latestGroupMemberList.append(members[0])
                for memb in memberList:
                    gms = GroupMember.objects.filter(group_id=latest_group.pk, member_id=memb.pk)
                    if gms.count() == 0: memberList2.append(memb)
            else:
                memberList2 = memberList
            gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
            recents = []
            for gc in gcs:
                gs = Group.objects.filter(id=gc.group_id)
                if gs.count() > 0:
                    group = gs[0]
                    group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                    recents.append(group)

            return render(request, 'motherwise/groups.html', {'members':memberList2, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
        else:
            return redirect('/manager/home')

    else:
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        memberIdList = []
        try:
            memberIdList = request.session['selected_member_list']
        except KeyError:
            print('No key')

        memberList = []
        for member_id in memberIdList:
            members = Member.objects.filter(id=member_id)
            if members.count() > 0:
                member = members[0]
                memberList.append(member)

        contacts = update_admin_contact(admin, "")

        if len(memberList) == 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'The members don\'t exist.'})

        if len(memberList) > 0:
            groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
            for group in groups:
                group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
            latestGroupMemberList = []
            latest_group = None
            if groups.count() > 0:
                latest_group = groups[0]
                gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        latestGroupMemberList.append(members[0])
            gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
            recents = []
            for gc in gcs:
                gs = Group.objects.filter(id=gc.group_id)
                if gs.count() > 0:
                    group = gs[0]
                    group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                    recents.append(group)
            return render(request, 'motherwise/groups.html', {'members':memberList, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
        else:
            return redirect('/manager/home')


def new_notis(request):
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    unreadNotiList = []
    notis = Received.objects.filter(member_id=admin.pk)
    for noti in notis:
        nfs = Notification.objects.filter(id=noti.noti_id)
        if nfs.count() > 0:
            notification = nfs[0]
            if notification.status == '':
                unreadNotiList.append(notification)

    return HttpResponse(len(unreadNotiList))



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def send_reply_message(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '1')
        noti_id = request.POST.get('noti_id', '1')
        message = request.POST.get('message', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        members = Member.objects.filter(id=int(member_id))
        if members.count() > 0:
            member = members[0]

            # title = 'You\'ve received a message from MotherWise Community'
            # subject = 'From MotherWise Community'
            # msg = 'Dear ' + member.name + ',<br><br>'
            # msg = msg + message

            # from_email = 'motherwisecolorado@gmail.com'
            # to_emails = []
            # to_emails.append(member.email)
            # send_mail_message(from_email, to_emails, title, subject, msg)

            msg = member.name + ', You\'ve received a reply message in the Nest.\nThe message is as following:\n' + message
            msg2 = member.name + ', has recibido un mensaje de respuesta en el Nest.\nel mensaje es el siguiente:\n' + message

            msg = msg + '\n\n' + msg2

            notification = Notification()
            notification.member_id = member.pk
            notification.sender_id = admin.pk
            notification.message = msg
            notification.notified_time = str(int(round(time.time() * 1000)))
            notification.save()

            rcv = Received()
            rcv.member_id = member.pk
            rcv.sender_id = admin.pk
            rcv.noti_id = notification.pk
            rcv.save()

            snt = Sent()
            snt.member_id = member.pk
            snt.sender_id = admin.pk
            snt.noti_id = notification.pk
            snt.save()

            replieds = Replied.objects.filter(noti_id=noti_id)
            if replieds.count() == 0:
                nfs = Notification.objects.filter(id=noti_id)
                if nfs.count() > 0:
                    noti = nfs[0]
                    replied = Replied()
                    replied.root_id = noti.pk
                    replied.noti_id = noti.pk
                    replied.save()

                    replied = Replied()
                    replied.root_id = noti.pk
                    replied.noti_id = notification.pk
                    replied.save()
            else:
                repl = replieds[0]
                replied = Replied()
                replied.root_id = repl.root_id
                replied.noti_id = notification.pk
                replied.save()

            ##########################################################################################################################################################################

            db = firebase.database()
            data = {
                "msg": msg,
                "date":str(int(round(time.time() * 1000))),
                "sender_id": str(admin.pk),
                "sender_name": admin.name,
                "sender_email": admin.email,
                "sender_photo": admin.photo_url,
                "role": "admin",
                "type": "message",
                "id": str(notification.pk),
                "mes_id": str(notification.pk)
            }

            db.child("notify").child(str(member.pk)).push(data)
            db.child("notify2").child(str(member.pk)).push(data)

            sendFCMPushNotification(member.pk, admin.pk, msg)

            #################################################################################################################################################################################

            if member.playerID != '':
                playerIDList = []
                playerIDList.append(member.playerID)
                msg = member.name + ', You\'ve received a reply message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
                url = '/mothers/notifications?noti_id=' + str(notification.pk)
                send_push(playerIDList, msg, url)

            return redirect('/manager/notifications/')
        else:
            return render(request, 'motherwise/result.html',
                          {'response': 'This member doesn\'t exist.'})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def send_member_message(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '1')
        message = request.POST.get('message', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        members = Member.objects.filter(id=int(member_id))
        if members.count() > 0:
            member = members[0]

            title = 'You\'ve received a message in the Nest'
            subject = 'MotherWise Community (Comunidad MotherWise: el Nest)'
            msg = 'Dear ' + member.name + ', You\'ve received a message in the Nest. The message is as following:<br><br>'
            msg = msg + message

            title2 = 'has recibido un mensaje en el Nest.'
            msg2 = member.name + ', has recibido un mensaje en el Nest. el mensaje es el siguiente:<br><br>'
            msg2 = msg2 + message

            from_email = admin.email
            to_emails = []
            to_emails.append(member.email)
            send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

            msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
            msg2 = member.name + ', has recibido un mensaje en el Nest.\nel mensaje es el siguiente:\n' + message

            msg = msg + '\n\n' + msg2

            notification = Notification()
            notification.member_id = member.pk
            notification.sender_id = admin.pk
            notification.message = msg
            notification.notified_time = str(int(round(time.time() * 1000)))
            notification.save()

            rcv = Received()
            rcv.member_id = member.pk
            rcv.sender_id = admin.pk
            rcv.noti_id = notification.pk
            rcv.save()

            snt = Sent()
            snt.member_id = member.pk
            snt.sender_id = admin.pk
            snt.noti_id = notification.pk
            snt.save()

            ##########################################################################################################################################################################

            db = firebase.database()
            data = {
                "msg": message,
                "date":str(int(round(time.time() * 1000))),
                "sender_id": str(admin.pk),
                "sender_name": admin.name,
                "sender_email": admin.email,
                "sender_photo": admin.photo_url,
                "role": "admin",
                "type": "message",
                "id": str(notification.pk),
                "mes_id": str(notification.pk)
            }

            db.child("notify").child(str(member.pk)).push(data)
            db.child("notify2").child(str(member.pk)).push(data)

            sendFCMPushNotification(member.pk, admin.pk, message)

            #################################################################################################################################################################################

            if member.playerID != '':
                playerIDList = []
                playerIDList.append(member.playerID)
                msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
                url = '/mothers/notifications?noti_id=' + str(notification.pk)
                send_push(playerIDList, msg, url)

            return redirect('/manager/notifications/')
        else:
            return render(request, 'motherwise/result.html',
                          {'response': 'This member doesn\'t exist.'})



def to_conferences(request):

    import datetime

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    c = Cohort.objects.filter(admin_id=admin.pk).first()
    cohorts = []
    if c is not None:
        if c.cohorts != '': cohorts = c.cohorts.split(',')

    confs = Conference.objects.filter(member_id=admin.pk).order_by('-id')
    confList = []
    for conf in confs:
        conf.created_time = datetime.datetime.fromtimestamp(float(int(conf.created_time)/1000)).strftime("%b %d, %Y %H:%M")
        if conf.event_time != '': conf.event_time = datetime.datetime.fromtimestamp(float(int(conf.event_time)/1000)).strftime("%b %d, %Y %H:%M")
        if int(conf.group_id) > 0:
            group = Group.objects.filter(id=conf.group_id).first()
            if group is not None:
                data={
                    'conf':conf,
                    'group': group
                }
                confList.append(data)
        else:
            data={
                'conf':conf,
                'group': None
            }
            confList.append(data)

    groups = Group.objects.filter(member_id=admin.pk).order_by('-id')

    return render(request, 'motherwise/conferences.html', {'conf_opt':'new_conference', 'confs':confList, 'cohorts':cohorts, 'groups':groups})




# from twilio.rest import Client

# def sendSMS(to_phone, msg):

#     # Your Account Sid and Auth Token from twilio.com/console
#     # DANGER! This is insecure. See http://twil.io/secure
#     account_sid = 'ACa84d7b1bddaec4ba6465060ae44fb2f3'
#     auth_token = 'bfc5cdda6bf320a153116fd80b2a9b7a'
#     client = Client(account_sid, auth_token)

#     message = client.messages \
#                     .create(
#                          body=msg,
#                          from_='+17206795056',
#                          to=to_phone
#                      )

#     print(message.sid)

#     return to_phone


# def sms_test(request):
#     to_phone = sendSMS('+18438161828', '12345')         #  18438161828
#     return HttpResponse('SMS sent to ' + to_phone)


def noti_detail(request):
    import datetime

    noti_id = request.GET['noti_id']
    opt = request.GET['opt']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    list = []
    sender0 = None

    replieds = Replied.objects.filter(noti_id=noti_id)
    if replieds.count() > 0:
        repl = replieds[0]
        repls = Replied.objects.filter(root_id=repl.root_id)
        for repl in repls:
            notis = Notification.objects.filter(id=repl.noti_id)
            if notis.count() > 0:
                noti = notis[0]
                date = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                members = Member.objects.filter(id=noti.sender_id)
                if members.count() > 0:
                    sender = members[0]
                    if sender.pk != admin.pk:
                        sender0 = sender
                    data = {
                        'sender':sender,
                        'noti': noti,
                        'date':date
                    }

                    list.append(data)

    else:
        notis = Notification.objects.filter(id=noti_id)
        if notis.count() > 0:
            noti = notis[0]
            date = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
            members = Member.objects.filter(id=noti.sender_id)
            if members.count() > 0:
                sender = members[0]
                sender0 = sender
                data = {
                    'sender':sender,
                    'noti': noti,
                    'date':date
                }

                list.append(data)

    return render(request, 'motherwise/noti_detail.html', {'notid':noti_id, 'me':admin, 'sender':sender0, 'list':list, 'opt':opt})



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def notify_group_chat(request):
    if request.method == 'POST':

        message = request.POST.get('message', '')
        cohort = request.POST.get('cohort', '')
        groupid = request.POST.get('groupid', '')
        ids = request.POST.getlist('users[]')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        if groupid != '':
            groups = Group.objects.filter(id=int(groupid))
            if groups.count() == 0:
                return redirect('/manager/home')

            group = groups[0]

            for member_id in ids:
                members = Member.objects.filter(id=int(member_id))
                if members.count() > 0:
                    member = members[0]

                    title = 'MotherWise Community: The Nest'
                    subject = 'You\'ve received a community message from (has recibido un mensaje de la comunidad de)' + group.name
                    msg = 'Dear ' + member.name + ', You\'ve received a community message from manager in ' + group.name + '. The message is as following:<br><br>'
                    msg = msg + message + '<br><br>'
                    msg = msg + '<a href=\'' + settings.URL + '/mothers/open_group_chat?group_id=' + groupid + '\' target=\'_blank\'>Connect the community to view message</a>'

                    title2 = 'Comunidad MotherWise: el Nest'
                    msg2 = member.name + ', has recibido un mensaje de la comunidad del administrador en ' + group.name + '. el mensaje es el siguiente:<br><br>'
                    msg2 = msg2 + message + '<br><br>'
                    msg2 = msg2 + '<a href=\'' + settings.URL + '/mothers/open_group_chat?group_id=' + groupid + '\' target=\'_blank\'>conectar la comunidad para ver el mensaje</a>'

                    from_email = admin.email
                    to_emails = []
                    to_emails.append(member.email)
                    send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                    msg = member.name + ', You\'ve received a community message from manager in ' + group.name + '. The message is as following:\n\n'
                    msg = msg + message + '\n\n'
                    msg = msg + 'Click on this link to view the message: ' + settings.URL + '/mothers/open_group_chat?group_id=' + groupid

                    msg2 = member.name + ', has recibido un mensaje de la comunidad del administrador en ' + group.name + '. el mensaje es el siguiente:\n\n'
                    msg2 = msg2 + message + '\n\n'
                    msg2 = msg2 + 'haga clic en este enlace para ver el mensaje: ' + settings.URL + '/mothers/open_group_chat?group_id=' + groupid

                    msg = msg + '\n\n' + msg2

                    notification = Notification()
                    notification.member_id = member.pk
                    notification.sender_id = admin.pk
                    notification.message = msg
                    notification.notified_time = str(int(round(time.time() * 1000)))
                    notification.save()

                    rcv = Received()
                    rcv.member_id = member.pk
                    rcv.sender_id = admin.pk
                    rcv.noti_id = notification.pk
                    rcv.save()

                    snt = Sent()
                    snt.member_id = member.pk
                    snt.sender_id = admin.pk
                    snt.noti_id = notification.pk
                    snt.save()

                    ##########################################################################################################################################################################

                    db = firebase.database()
                    data = {
                        "msg": msg,
                        "date":str(int(round(time.time() * 1000))),
                        "sender_id": str(admin.pk),
                        "sender_name": admin.name,
                        "sender_email": admin.email,
                        "sender_photo": admin.photo_url,
                        "role": "admin",
                        "type": "group_chat",
                        "id": str(group.pk),
                        "mes_id": str(notification.pk)
                    }

                    db.child("notify").child(str(member.pk)).push(data)
                    db.child("notify2").child(str(member.pk)).push(data)

                    sendFCMPushNotification(member.pk, admin.pk, msg)

                    #################################################################################################################################################################################

                    if member.playerID != '':
                        playerIDList = []
                        playerIDList.append(member.playerID)
                        url = '/mothers/notifications?noti_id=' + str(notification.pk)
                        send_push(playerIDList, msg, url)

        elif cohort != '':

            for member_id in ids:
                members = Member.objects.filter(id=int(member_id))
                if members.count() > 0:
                    member = members[0]

                    title = 'MotherWise Community: The Nest'
                    subject = 'You\'ve received a group message from ' + cohort
                    msg = 'Dear ' + member.name + ', You\'ve received a group message from manager in ' + cohort + '. The message is as following:<br><br>'
                    msg = msg + message + '<br><br>'
                    msg = msg + '<a href=\'' + settings.URL + '/mothers/open_cohort_chat?cohort=' + cohort + '\' target=\'_blank\'>Connect the group to view message</a>'

                    title2 = 'Comunidad MotherWise: el Nest'
                    msg2 = member.name + ', has recibido un mensaje de la comunidad del administrador en ' + cohort + '. el mensaje es el siguiente:<br><br>'
                    msg2 = msg2 + message + '<br><br>'
                    msg2 = msg2 + '<a href=\'' + settings.URL + '/mothers/open_cohort_chat?cohort=' + cohort + '\' target=\'_blank\'>conectar el grupo para ver el mensaje</a>'

                    from_email = admin.email
                    to_emails = []
                    to_emails.append(member.email)
                    send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                    msg = member.name + ', You\'ve received a group message from manager in ' + cohort + '. The message is as following:\n\n'
                    msg = msg + message + '\n\n'
                    msg = msg + 'Click on this link to view the message: ' + settings.URL + '/mothers/open_cohort_chat?cohort=' + cohort

                    msg2 = member.name + ', has recibido un mensaje de la comunidad del administrador en ' + cohort + '. el mensaje es el siguiente:\n\n'
                    msg2 = msg2 + message + '\n\n'
                    msg2 = msg2 + 'haga clic en este enlace para ver el mensaje: ' + settings.URL + '/mothers/open_cohort_chat?cohort=' + cohort

                    msg = msg + '\n\n' + msg2

                    notification = Notification()
                    notification.member_id = member.pk
                    notification.sender_id = admin.pk
                    notification.message = msg
                    notification.notified_time = str(int(round(time.time() * 1000)))
                    notification.save()

                    rcv = Received()
                    rcv.member_id = member.pk
                    rcv.sender_id = admin.pk
                    rcv.noti_id = notification.pk
                    rcv.save()

                    snt = Sent()
                    snt.member_id = member.pk
                    snt.sender_id = admin.pk
                    snt.noti_id = notification.pk
                    snt.save()

                    ##########################################################################################################################################################################

                    db = firebase.database()
                    data = {
                        "msg": msg,
                        "date":str(int(round(time.time() * 1000))),
                        "sender_id": str(admin.pk),
                        "sender_name": admin.name,
                        "sender_email": admin.email,
                        "sender_photo": admin.photo_url,
                        "role": "admin",
                        "type": "cohort_chat",
                        "id": cohort,
                        "mes_id": str(notification.pk)
                    }

                    db.child("notify").child(str(member.pk)).push(data)
                    db.child("notify2").child(str(member.pk)).push(data)

                    sendFCMPushNotification(member.pk, admin.pk, msg)

                    #################################################################################################################################################################################

                    if member.playerID != '':
                        playerIDList = []
                        playerIDList.append(member.playerID)
                        url = '/mothers/notifications?noti_id=' + str(notification.pk)
                        send_push(playerIDList, msg, url)

        return HttpResponse('success')



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def sendfcmpush(request):
    if request.method == 'POST':

        sender_id = request.POST.get('sender_id', '1')
        receiver_id = request.POST.get('receiver_id', '1')
        message = request.POST.get('message', '')

        sendFCMPushNotification(receiver_id, sender_id, message)

        senders = Member.objects.filter(id=sender_id)
        if senders.count() == 0:
            resp = {'result_code': '1'}
            return HttpResponse(json.dumps(resp))

        sender = senders[0]

        members = Member.objects.filter(id=receiver_id)
        if members.count() > 0:
            member = members[0]
            if member.playerID != '':
                playerIDList = []
                playerIDList.append(member.playerID)
                url = '/users/to_private_chat?member_id=' + str(sender.pk)
                if member.cohort == 'admin':
                    url = '/group_private_chat?email=' + sender.email
                msg = member.name + ', You\'ve received a message from ' + sender.name + '.\nThe message is as following:\n' + message
                msg2 = member.name + ', has recibido un mensaje de ' + sender.name + '.\nel mensaje es el siguiente:\n' + message
                msg = msg + '\n\n' + msg2
                send_push(playerIDList, msg, url)

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp))



def sendFCMPushNotification(member_id, sender_id, notiText):
    members = Member.objects.filter(id=member_id)
    if members.count() > 0:
        member = members[0]
        message_title = 'VaCay User'
        if int(sender_id) > 0:
            senders = Member.objects.filter(id=sender_id)
            if senders.count() > 0:
                sender = senders[0]
                message_title = sender.name
        path_to_fcm = "https://fcm.googleapis.com"
        server_key = settings.FCM_LEGACY_SERVER_KEY
        reg_id = member.fcm_token #quick and dirty way to get that ONE fcmId from table
        if reg_id != '':
            message_body = notiText
            result = FCMNotification(api_key=server_key).notify_single_device(registration_id=reg_id, message_title=message_title, message_body=message_body, sound = 'ping.aiff', badge = 1)



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def verifypwreport(request):
    if request.method == 'POST':
        password = request.POST.get('password', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        if password == admin.password:
            return HttpResponse('0')
        else:
            return HttpResponse('1')


def reports(request):
    import datetime
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    reports = Report.objects.all()
    reportList = []
    for report in reports:
        members = Member.objects.filter(id=int(report.member_id))
        if members.count() > 0:
            member = members[0]
            member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
            if int(member.admin_id) == admin.pk:
                reporters = Member.objects.filter(id=int(report.reporter_id))
                if reporters.count() > 0:
                    reporter = reporters[0]
                    data = {
                        'id':report.pk,
                        'member':member,
                        'reporter':reporter,
                        'message':report.message,
                        'date_time':datetime.datetime.fromtimestamp(float(int(report.reported_time)/1000)).strftime("%b %d, %Y %H:%M"),
                        'status':report.status
                    }
                    reportList.append(data)
    return render(request, 'motherwise/reports.html', {'reports':reportList})


def delreport(request):
    report_id = request.GET['report_id']
    reports = Report.objects.filter(id=int(report_id))
    if reports.count() > 0:
        report = reports[0]
        report.delete()
        return redirect('/reports')
    else:
        return render(request, 'motherwise/result.html', {'response':'Something is wrong. The report doesn\'t exist.'})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def warning_message(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '1')
        message = request.POST.get('message', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        members = Member.objects.filter(id=int(member_id))
        if members.count() > 0:
            member = members[0]

            title = 'You\'ve received a warning message in the Nest'
            subject = 'MotherWise Community (Comunidad MotherWise: el Nest)'
            msg = member.name + ', You\'ve received a warning message in the Nest. The message is as following:<br><br>'
            msg = msg + message

            title2 = 'has recibido un mensaje de advertencia en el Nest:'
            msg2 = member.name + ', has recibido un mensaje de advertencia en el Nest. el mensaje es el siguiente:<br><br>'
            msg2 = msg2 + message

            from_email = admin.email
            to_emails = []
            to_emails.append(member.email)
            send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

            msg = member.name + ', You\'ve received a warning message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
            msg2 = member.name + ', has recibido un mensaje de advertencia en el Nest.\nel mensaje es el siguiente:\n' + message

            msg = msg + '\n\n' + msg2

            notification = Notification()
            notification.member_id = member.pk
            notification.sender_id = admin.pk
            notification.message = msg
            notification.notified_time = str(int(round(time.time() * 1000)))
            notification.save()

            rcv = Received()
            rcv.member_id = member.pk
            rcv.sender_id = admin.pk
            rcv.noti_id = notification.pk
            rcv.save()

            snt = Sent()
            snt.member_id = member.pk
            snt.sender_id = admin.pk
            snt.noti_id = notification.pk
            snt.save()

            ##########################################################################################################################################################################

            db = firebase.database()
            data = {
                "msg": message,
                "date":str(int(round(time.time() * 1000))),
                "sender_id": str(admin.pk),
                "sender_name": admin.name,
                "sender_email": admin.email,
                "sender_photo": admin.photo_url,
                "role": "admin",
                "type": "message",
                "id": str(notification.pk),
                "mes_id": str(notification.pk)
            }

            db.child("notify").child(str(member.pk)).push(data)
            db.child("notify2").child(str(member.pk)).push(data)

            sendFCMPushNotification(member.pk, admin.pk, message)

            #################################################################################################################################################################################

            if member.playerID != '':
                playerIDList = []
                playerIDList.append(member.playerID)
                msg = member.name + ', You\'ve received a warning message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
                msg2 = member.name + ', has recibido un mensaje de advertencia en el Nest.\nel mensaje es el siguiente:\n' + message
                msg = msg + '\n\n' + msg2
                url = '/mothers/notifications?noti_id=' + str(notification.pk)
                send_push(playerIDList, msg, url)

            return HttpResponse('0')
        else:
            return HttpResponse('1')




def reinvite_member(request):
    member_id = request.GET['member_id']
    members = Member.objects.filter(id=member_id)
    if members.count() > 0:
        member = members.first()
        name = member.name
        email = member.email
        phone_number = member.phone_number
        cohort = member.cohort

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return HttpResponse('no_admin_auth')
        except KeyError:
            print('no session')
            return HttpResponse('no_admin_auth')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        groupText = ''
        if member.cohort != '':
            groupText = '<br>Group: ' + member.cohort

        title = 'Invitation for MotherWise Community: The Nest'
        subject = 'MotherWise Community: The Nest'
        message = 'Dear ' + member.name + ',<br><br>Welcome to \"The Nest\": MotherWise\'s virtual community!<br><br>The Nest is an opportunity to connect and reconnect with other MotherWise families.<br>'
        message = message + 'You can post articles, share pregnancy and new baby tips, watch videos, and chat directly with other moms. You\'ll also stay up-to-date on all the new programs and special events MotherWise has to offer!<br><br>'
        message = message + settings.URL + '/nest/mothers' + '<br><br>We are providing you with your initial login information as follows:<br><br>'
        message = message + 'E-mail: ' + member.email + ' (your email)<br>Password: ' + member.password + groupText + '<br><br>'

        message = message + '***By signing up to The Nest, you are agreeing to not engage in any type of: ***<br>'
        message = message + '        · hate speech<br>'
        message = message + '        · cyberbullying<br>'
        message = message + '        · solicitation and/or selling of goods or services<br>'
        message = message + '        · posting content inappropriate for our diverse community including but not limited to political<br>'
        message = message + '    or religious views<br><br>'
        message = message + 'We want The Nest to be a safe place for support and inspiration. Help us foster this community and please respect everyone on The Nest.<br><br>'
        message = message + 'Please watch this video to see how to login: https://vimeo.com/430742850<br><br>'
        message = message + 'If you have any question, please contact us:<br><br>'

        message = message + '   E-mail: ' + 'motherwisecolorado@gmail.com' + '<br>   Phone number: ' + '720-504-4624<br><br>'
        message = message + '<a href=\'' + settings.URL + '/nest/mothers' + '\' target=\'_blank\'>Join website</a><br><br>'
        message = message + 'Sincerely<br><br>MotherWise Team'

        message = message + '<br><br>'

        groupText2 = ''
        if member.cohort != '':
            groupText2 = '<br>Grupo: ' + member.cohort

        title2 = 'Invitación para la comunidad de MotherWise: El Nido'
        message2 = 'Querida ' + member.name + ',<br><br>¡Bienvenida al \"Nido\": la comunidad virtual de MotherWise!<br><br>El Nido es una oportunidad para conectarse y reconectarse con otras madres de MotherWise.<br>'
        message2 = message2 + 'Puede publicar artículos, compartir consejos sobre embarazo y nuevos bebés, ver videos y chatear directamente con otras madres. ¡También se mantendrá al tanto sobre todos los nuevos programas y eventos especiales que MotherWise tiene para ofrecer!<br><br>'
        message2 = message2 + settings.URL + '/nest/mothers' + '<br><br>Le proporcionamos su información de inicio de la siguiente manera:<br><br>'
        message2 = message2 + 'Correo electrónico: ' + member.email + ' (Tu correo electrónico)<br>Contraseña: ' + member.password + groupText2 + '<br><br>'

        message2 = message2 + '***Al suscribirse al Nido, acepta no participar en ningún tipo de: ***<br>'
        message2 = message2 + '        · El discurso del odio<br>'
        message2 = message2 + '        · Ciberacoso<br>'
        message2 = message2 + '        · Solicitud y/o venta de bienes o servicios<br>'
        message2 = message2 + '        · Publicar contenido inapropiado para nuestra diversa comunidad, incluidos, entre otros, puntos de vista políticos o religiosos<br><br>'
        message2 = message2 + 'Queremos que El Nido sea un lugar seguro para apoyo e inspiración. Por favor ayúdanos a fomentar esta comunidad y por favor respeta a todos en El Nido.<br><br>'
        message2 = message2 + 'Mire este video para ver cómo iniciar sesión: https://vimeo.com/430742850<br><br>'
        message2 = message2 + 'Si usted tiene cualquier pregunta, por favor póngase en contacto con nosotros:<br><br>'

        message2 = message2 + '   Correo electrónico: ' + 'motherwisecolorado@gmail.com' + '<br>   Número de teléfono: ' + '720-504-4624<br><br>'
        message2 = message2 + '<a href=\'' + settings.URL + '/nest/mothers' + '\' target=\'_blank\'>Unirse al sitio web</a><br><br>'
        message2 = message2 + 'Sinceramente,<br><br>el Equipo de MotherWise'

        from_email = admin.email
        to_emails = []
        to_emails.append(member.email)
        send_mail_message0(from_email, to_emails, title, subject, message, title2, message2)

        return HttpResponse('success')

    return HttpResponse('no_auth')




def toplevelsetup(request):
    post_id = request.GET['post_id']
    posts = Post.objects.filter(id=post_id, sch_status='')
    if posts.count() > 0:
        post = posts.first()
        if not 'top' in post.status:
            post.status += 'top'
        else:
            post.status = post.status.replace('top','')
        post.save()
    return redirect('/manager/posts')





def now():
    from datetime import datetime
    return datetime.now()


def always_on(request):

    year = now().year
    month = now().month
    day = now().day
    hour = now().hour
    minute = now().minute

    list = []

    import datetime
    posts = Post.objects.filter(~Q(scheduled_time=''))
    for post in posts:
        splits = post.scheduled_time.split('-')
        pyear = int(splits[0])
        pmonth = int(splits[1])
        pday = int(splits[2])
        phour = int(splits[3])
        pminute = int(splits[4])

        sxxx = datetime.datetime(pyear, pmonth, pday, phour, pminute, 1)
        nxxx = datetime.datetime(year, month, day, hour, minute, 1)
        duration = nxxx - sxxx

        list.append(duration.total_seconds())

        if duration.total_seconds() >= 0:

            if post.notified_members != '' and post.sch_status != '':
                ids = post.notified_members.split(',')
                for m_id in ids:
                    members = Member.objects.filter(id=int(m_id))
                    if members.count() > 0:
                        member = members.first()
                        admin = Member.objects.get(id=member.admin_id)

                        title = 'MotherWise Community: The Nest'
                        subject = 'You\'ve received a post in the Nest (has recibido una publicación en el Nest)'
                        msg = 'Dear ' + member.name + ', You\'ve received a post from MotherWise Community: The Nest.<br><br>'
                        msg = msg + '<a href=\'' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk) + '\' target=\'_blank\'>View the post</a>'

                        title2 = 'comunidad MotherWise: el Nest'
                        msg2 = member.name + ', has recibido una publicación en el Nest.<br><br>'
                        msg2 = msg2 + '<a href=\'' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk) + '\' target=\'_blank\'>ver la publicación</a>'

                        from_email = admin.email
                        to_emails = []
                        to_emails.append(member.email)
                        send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                        msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\n\n'
                        msg = msg + 'Click on this link to view the post: ' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk)

                        msg2 = member.name + ', has recibido una publicación en el Nest.\n\n'
                        msg2 = msg2 + 'haga clic en este enlace para ver la publicación: ' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk)

                        msg = msg + '\n\n' + msg2

                        notification = Notification()
                        notification.member_id = member.pk
                        notification.sender_id = admin.pk
                        notification.message = msg
                        notification.notified_time = str(int(round(time.time() * 1000)))
                        notification.save()

                        rcv = Received()
                        rcv.member_id = member.pk
                        rcv.sender_id = admin.pk
                        rcv.noti_id = notification.pk
                        rcv.save()

                        snt = Sent()
                        snt.member_id = member.pk
                        snt.sender_id = admin.pk
                        snt.noti_id = notification.pk
                        snt.save()

                        ##########################################################################################################################################################################

                        db = firebase.database()
                        data = {
                            "msg": msg,
                            "date":str(int(round(time.time() * 1000))),
                            "sender_id": str(admin.pk),
                            "sender_name": admin.name,
                            "sender_email": admin.email,
                            "sender_photo": admin.photo_url,
                            "role": "admin",
                            "type": "post",
                            "id": str(post.pk),
                            "mes_id": str(notification.pk)
                        }

                        db.child("notify").child(str(member.pk)).push(data)
                        db.child("notify2").child(str(member.pk)).push(data)

                        sendFCMPushNotification(member.pk, admin.pk, msg)

                        #################################################################################################################################################################################

                        if member.playerID != '':
                            playerIDList = []
                            playerIDList.append(member.playerID)
                            url = '/mothers/notifications?noti_id=' + str(notification.pk)
                            if member.cohort == 'admin':
                                url = '/manager/notifications?noti_id=' + str(notification.pk)
                            send_push(playerIDList, msg, url)

            if post.sch_status != '':
                post.sch_status = ''
                post.save()

    return HttpResponse(str(list) + ' /// ' + str(now()))




def openbroadcast(request):
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return HttpResponse('no_admin_auth')
    except KeyError:
        print('no session')
        return HttpResponse('no_admin_auth')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)
    users = Member.objects.filter(~Q(id=admin.pk)).order_by('-id')
    return render(request, 'motherwise/broadcast.html', {'users':users})


@api_view(['GET', 'POST'])
def broadcast(request):
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return HttpResponse('no_admin_auth')
    except KeyError:
        print('no session')
        return HttpResponse('no_admin_auth')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    if request.method == 'POST':
        message = request.POST.get('message', '')

        is_file = False

        fname = None
        fread = None
        fcontenttype = None

        try:
            f = request.FILES['file']
            if f.size > 5242880:
                is_file = False
            else:
                fname = f.name
                fread = f.read()
                fcontenttype = f.content_type
                is_file = True
        except MultiValueDictKeyError:
            is_file = False

        members = Member.objects.filter(~Q(id=admin.pk)).order_by('-id')

        for member in members:
            title = 'MotherWise Community: The Nest'
            subject = 'You\'ve received a message from the Nest (Has recibido un mensaje del Nest)'

            from_email = admin.email
            to_emails = []
            if member.registered_time != '' and member.notice_excluded == '':
                to_emails.append(member.email)

                try:
                    if is_file == False: send_mail_message(from_email, to_emails, title, subject, message)
                    else: send_mail_message_with_file(from_email, to_emails, title, subject, message, fname, fread, fcontenttype)
                except: pass

                try: sendFCMPushNotification(member.pk, admin.pk, message)
                except: pass

                notification = Notification()
                notification.member_id = member.pk
                notification.sender_id = admin.pk
                notification.message = message
                notification.notified_time = str(int(round(time.time() * 1000)))
                notification.save()

                if member.playerID != '':
                    try:
                        playerIDList = []
                        playerIDList.append(member.playerID)
                        msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\nThe message is as following:\n' + message
                        msg2 = member.name + ', has recibido un mensaje de Nest.\nel mensaje es el siguiente:\n' + message
                        msg = msg + '\n\n' + msg2
                        url = '/mothers/notifications?noti_id=' + str(notification.pk)
                        send_push(playerIDList, msg, url)
                    except: pass

        return HttpResponse('success')



@api_view(['POST','GET'])
def bexcludedemailsave(request):
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return HttpResponse('no_admin_auth')
    except KeyError:
        print('no session')
        return HttpResponse('no_admin_auth')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    if request.method == 'POST':
        exusers = request.POST.getlist('exusers[]')

        members = Member.objects.filter(~Q(id=admin.pk)).order_by('-id')
        for member in members:
            if member.notice_excluded != '':
                member.notice_excluded = ''
                member.save()
        for uid in exusers:
            member = members.filter(id=uid).first()
            if member is not None:
                member.notice_excluded = 'yes'
                member.save()
        return HttpResponse(json.dumps({'result':'success', 'exusers':str(len(exusers))}))



def send_mail_message_with_file (from_email, to_emails, title, subject, message, fname, fread, fcontenttype):
    html =  """\
                <html>
                    <head></head>
                    <body>

                        <h2 style="margin-left:10px; color:#02839a;">{title}</h2>
                        <div style="font-size:14px; white-space: pre-line; word-wrap: break-word;">
                            {mes}
                        </div>
                    </body>
                </html>
            """
    html = html.format(title=title, mes=message)

    mail = EmailMultiAlternatives(subject, '', from_email, to_emails)
    mail.attach_alternative(html, "text/html")
    mail.attach(fname, fread, fcontenttype)
    mail.send(fail_silently=False)



#####################################################################################################################################

def tonewpost(request):
    import datetime
    try:
        if request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
    userList = []
    for user in users:
        if user.registered_time != '':
            user.username = '@' + user.email[0:user.email.find('@')]
            userList.append(user)

    categories = []
    pc = PostCategory.objects.filter(admin_id=admin.pk).first()
    if pc is not None:
        if pc.categories != '': categories = pc.categories.split(',')

    token = request.GET['token']

    comida = ''
    try: comida = request.GET['comida']
    except: pass

    return render(request, 'mothers/new_post.html', {'me':admin, 'users':userList, 'token':token, 'comida':comida, 'categories':categories, 'opt':'admin'})




@api_view(['GET', 'POST'])
def newpost(request):
    if request.method == 'POST':

        title = request.POST.get('title', '')
        category = request.POST.get('category', '')
        content = request.POST.get('content', '')
        scheduled_time = request.POST.get('scheduled_time', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return HttpResponse('error')
        except KeyError:
            print('no session')
            return HttpResponse('error')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        post = Post()
        post.member_id = admin.pk
        post.title = title
        post.category = category
        post.content = emoji.demojize(content)
        post.picture_url = ''
        post.comments = '0'
        post.likes = '0'
        post.loves = '0'
        post.haha = '0'
        post.wow = '0'
        post.sad = '0'
        post.angry = '0'
        post.reactions = '0'
        post.scheduled_time = scheduled_time
        post.posted_time = str(int(round(time.time() * 1000)))

        try:
            ids = request.POST.getlist('users[]')
            if len(ids) > 0: post.notified_members = ",".join(str(i) for i in ids)
        except KeyError:
            print('No key')

        try:
            ids = request.POST.get('selections','')
            if ids != '': post.notified_members = ids
        except KeyError:
            print('No key')

        if scheduled_time != '': post.sch_status = 'scheduled'
        post.save()

        createPostUrlPreview(post)

        fs = FileSystemStorage()
        i = 0
        for f in request.FILES.getlist('pictures'):
            i = i + 1
            filename = fs.save(f.name, f)
            uploaded_url = fs.url(filename)
            if i == 1:
                post.picture_url = settings.URL + uploaded_url
                post.save()
            postPicture = PostPicture()
            postPicture.post_id = post.pk
            postPicture.picture_url = settings.URL + uploaded_url
            postPicture.filename = filename
            postPicture.save()


        if post.scheduled_time == '':

            for member_id in ids:
                members = Member.objects.filter(id=int(member_id))
                if members.count() > 0:
                    member = members[0]

                    title = 'MotherWise Community: The Nest'
                    subject = 'You\'ve received a post in the Nest (has recibido una publicación en el Nest)'
                    msg = 'Dear ' + member.name + ', You\'ve received a post from MotherWise Community: The Nest.<br><br>'
                    msg = msg + '<a href=\'' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk) + '\' target=\'_blank\'>View the post</a>'

                    title2 = 'comunidad MotherWise: el Nest'
                    msg2 = member.name + ', has recibido una publicación en el Nest.<br><br>'
                    msg2 = msg2 + '<a href=\'' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk) + '\' target=\'_blank\'>ver la publicación</a>'

                    from_email = admin.email
                    to_emails = []
                    to_emails.append(member.email)
                    send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                    msg = member.name + ', You\'ve received a message from MotherWise Community: The Nest.\n\n'
                    msg = msg + 'Click on this link to view the post: ' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk)

                    msg2 = member.name + ', has recibido una publicación en el Nest.\n\n'
                    msg2 = msg2 + 'haga clic en este enlace para ver la publicación: ' + settings.URL + '/mothers/to_post?post_id=' + str(post.pk)

                    msg = msg + '\n\n' + msg2

                    notification = Notification()
                    notification.member_id = member.pk
                    notification.sender_id = admin.pk
                    notification.message = msg
                    notification.notified_time = str(int(round(time.time() * 1000)))
                    notification.save()

                    rcv = Received()
                    rcv.member_id = member.pk
                    rcv.sender_id = admin.pk
                    rcv.noti_id = notification.pk
                    rcv.save()

                    snt = Sent()
                    snt.member_id = member.pk
                    snt.sender_id = admin.pk
                    snt.noti_id = notification.pk
                    snt.save()

                    ##########################################################################################################################################################################

                    db = firebase.database()
                    data = {
                        "msg": msg,
                        "date":str(int(round(time.time() * 1000))),
                        "sender_id": str(admin.pk),
                        "sender_name": admin.name,
                        "sender_email": admin.email,
                        "sender_photo": admin.photo_url,
                        "role": "admin",
                        "type": "post",
                        "id": str(post.pk),
                        "mes_id": str(notification.pk)
                    }

                    db.child("notify").child(str(member.pk)).push(data)
                    db.child("notify2").child(str(member.pk)).push(data)

                    sendFCMPushNotification(member.pk, admin.pk, msg)

                    #################################################################################################################################################################################

                    if member.playerID != '':
                        playerIDList = []
                        playerIDList.append(member.playerID)
                        url = '/mothers/notifications?noti_id=' + str(notification.pk)
                        send_push(playerIDList, msg, url)

        return HttpResponse('success')



def analytics(request):

    from datetime import datetime

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return HttpResponse('error')
    except KeyError:
        print('no session')
        return HttpResponse('error')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(admin_id=admin.pk)
    groups = Cohort.objects.filter(admin_id=admin.pk).first().cohorts.split(',')
    gcnt = len(groups)
    groups1 = groups[:int(gcnt / 2)]
    groups2 = groups[int(gcnt / 2):gcnt]

    inviteds1 = [0] * len(groups1)
    activateds1 = [0] * len(groups1)
    group_posts1 = [0] * len(groups1)

    inviteds2 = [0] * len(groups2)
    activateds2 = [0] * len(groups2)
    group_posts2 = [0] * len(groups2)

    total_activs = 0
    activ_percentlist_bygroup = []

    monthly_activateds1 = [0,0,0,0,0,0,0,0,0,0,0,0]
    monthly_posts1 = [0,0,0,0,0,0,0,0,0,0,0,0]
    monthly_activateds2 = [0,0,0,0,0,0,0,0,0,0,0,0]
    monthly_posts2 = [0,0,0,0,0,0,0,0,0,0,0,0]
    monthly_activateds3 = [0,0,0,0,0,0,0,0,0,0,0,0]
    monthly_posts3 = [0,0,0,0,0,0,0,0,0,0,0,0]

    cities = []
    activateds_bycity = []
    activated_members_bycity = []

    this_year = datetime.fromtimestamp(time.time()).year
    last_year1 = this_year - 1
    last_year2 = this_year - 2

    for member in members:
        if member.cohort != 'admin' and member.cohort != '':
            if member.cohort in groups1:
                index_invited = groups1.index(member.cohort)
                invits = inviteds1[index_invited] + 1
                inviteds1[index_invited] = invits
                if member.registered_time != '' and int(member.registered_time) > 0:
                    index_activated = groups1.index(member.cohort)
                    activs = activateds1[index_activated] + 1
                    activateds1[index_activated] = activs
                    total_activs = total_activs + 1

                    posts = Post.objects.filter(member_id=member.pk)
                    for post in posts:
                        psts = group_posts1[index_activated] + 1
                        group_posts1[index_activated] = psts
            elif member.cohort in groups2:
                index_invited = groups2.index(member.cohort)
                invits = inviteds2[index_invited] + 1
                inviteds2[index_invited] = invits
                if member.registered_time != '' and int(member.registered_time) > 0:
                    index_activated = groups2.index(member.cohort)
                    activs = activateds2[index_activated] + 1
                    activateds2[index_activated] = activs
                    total_activs = total_activs + 1

                    posts = Post.objects.filter(member_id=member.pk)
                    for post in posts:
                        psts = group_posts2[index_activated] + 1
                        group_posts2[index_activated] = psts


        if member.registered_time != '' and int(member.registered_time) > 0:
            registered_date_obj = datetime.fromtimestamp(int(member.registered_time)/1000)
            if this_year == registered_date_obj.year:
                activs = monthly_activateds1[registered_date_obj.month - 1] + 1
                monthly_activateds1[registered_date_obj.month - 1] = activs
            if last_year1 == registered_date_obj.year:
                activs = monthly_activateds2[registered_date_obj.month - 1] + 1
                monthly_activateds2[registered_date_obj.month - 1] = activs
            if last_year2 == registered_date_obj.year:
                activs = monthly_activateds3[registered_date_obj.month - 1] + 1
                monthly_activateds3[registered_date_obj.month - 1] = activs
            if member.city.replace('\'','').strip() not in cities:
                cities.append(member.city.replace('\'','').strip())
                activateds_bycity.append(0)
                activated_members_bycity.append([])
            activs = activateds_bycity[cities.index(member.city.replace('\'','').strip())] + 1
            activateds_bycity[cities.index(member.city.replace('\'','').strip())] = activs
            activated_members_bycity[cities.index(member.city.replace('\'','').strip())].append(member)

            posts = Post.objects.filter(member_id=member.pk)
            for post in posts:
                posted_date_obj = datetime.fromtimestamp(int(post.posted_time)/1000)
                if this_year == posted_date_obj.year:
                    psts = monthly_posts1[posted_date_obj.month - 1] + 1
                    monthly_posts1[posted_date_obj.month - 1] = psts
                if last_year1 == posted_date_obj.year:
                    psts = monthly_posts2[posted_date_obj.month - 1] + 1
                    monthly_posts2[posted_date_obj.month - 1] = psts
                if last_year2 == posted_date_obj.year:
                    psts = monthly_posts3[posted_date_obj.month - 1] + 1
                    monthly_posts3[posted_date_obj.month - 1] = psts


    cityActivsList = []
    cityActivsChartWidth = 0
    for i in range(0, len(cities)):
        data = {
            'city': cities[i],
            'activsval': activateds_bycity[i],
            'activmembers': activated_members_bycity[i]
        }
        cityActivsList.append(data)
        cityActivsChartWidth = cityActivsChartWidth + 50

    activateds = activateds1 + activateds2

    for i in range(0, len(activateds)):
        percentval = round(activateds[i] * 100 / total_activs, 2)
        data = {
            'group': groups[i],
            'activ_percent': percentval
        }
        activ_percentlist_bygroup.append(data)

    activ_percent = round(total_activs * 100 / members.count(), 2)
    activdata = {
        'percvalue': activ_percent,
        'title': 'Activated'
    }
    inactivdata = {
        'percvalue': round(100 - activ_percent, 2),
        'title': 'Inactivated'
    }

    communities = Group.objects.filter(member_id=admin.pk)
    comActivsList = []
    comPostsList = []
    for com in communities:
        gms = GroupMember.objects.filter(group_id=com.pk)
        member_count = gms.count()
        data = {
            'com': com.name,
            'activsval': str(member_count),
        }
        comActivsList.append(data)
        post_count = 0
        for gm in gms:
            post_count += Post.objects.filter(member_id=gm.member_id).count()
        data = {
            'com': com.name,
            'postsval': str(post_count),
        }
        comPostsList.append(data)

    context = {
        'this_year': str(this_year),
        'last_year1': str(last_year1),
        'last_year2': str(last_year2),
        'groups': groups,
        'groups1': groups1,
        'groups2': groups2,
        'activateds': activateds,
        'inviteds1': inviteds1,
        'activateds1': activateds1,
        'group_posts1': group_posts1,
        'inviteds2': inviteds2,
        'activateds2': activateds2,
        'group_posts2': group_posts2,
        'activ_percentlist_bygroup': activ_percentlist_bygroup,
        'total_activ_inactiv_percent': [activdata, inactivdata],
        'monthly_activateds1': monthly_activateds1,
        'monthly_posts1': monthly_posts1,
        'monthly_activateds2': monthly_activateds2,
        'monthly_posts2': monthly_posts2,
        'monthly_activateds3': monthly_activateds3,
        'monthly_posts3': monthly_posts3,
        'city_activateds': cityActivsList,
        'cityActivsChartWidth': cityActivsChartWidth,
        'com_activateds': comActivsList,
        'com_posts': comPostsList,
    }

    return render(request, 'motherwise/analytics.html', context)





































def open_translate(request):
    return render(request, 'motherwise/translate.html')


@api_view(['GET', 'POST'])
def process_translate(request):
    from googletrans import Translator
    if request.method == 'POST':
        input = request.POST.get('input', '')
        lang = request.POST.get('lang', '')

        translator = Translator()
        translation = translator.translate(input, dest=lang)
        return HttpResponse(translation.text)




def smtp_test(request):
    send_mail_message('motherwisecolorado@gmail.com', ['marquez07melissa@gmail.com'], 'From MotherWise Manager', 'Test message', 'Hello how are you?')
    return HttpResponse('success!')



def testtranslate(request):
    from googletrans import Translator
    sentences = [
        "Challenge",
        "Resources",
        "MotherWise Announcements",
        "Just for Fun",
    ]
    translator = Translator()
    translations = translator.translate(sentences, dest="es")
    arr = []
    for translation in translations:
        arr.append(translation.text)
    return HttpResponse(",".join(arr))








































